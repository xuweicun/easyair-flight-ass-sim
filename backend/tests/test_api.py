import os
import sqlite3
from io import BytesIO
from pathlib import Path

os.environ["DATABASE_URL"] = "sqlite+aiosqlite:///./test_flight_simulator.db"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402


def test_demo_api_supports_review_and_unassigned_export() -> None:
    database = Path("test_flight_simulator.db")
    if database.exists():
        database.unlink()

    with TestClient(app) as client:
        dashboard = client.get("/api/dashboard")
        assert dashboard.status_code == 200
        assert dashboard.json()["active_run"]["metrics"]["node_conservation"] is True

        groups = client.get("/api/groups").json()
        assert groups
        missing = next(group for group in groups if "MISSING_PLAN" in group["issue_tags"])

        detail = client.get(f"/api/groups/{missing['id']}")
        assert detail.status_code == 200
        assert detail.json()["nodes"]
        assert all(node["phase"] for node in detail.json()["nodes"])
        assert "related_segments" in detail.json()

        run_id = dashboard.json()["active_run"]["id"]
        node_anomalies = client.get(f"/api/runs/{run_id}/node-anomalies")
        assert node_anomalies.status_code == 200
        assert "affected_stands" in node_anomalies.json()["statistics"]
        anomaly_export = client.get(f"/api/runs/{run_id}/exports/node-anomalies.xlsx")
        assert anomaly_export.status_code == 200
        assert anomaly_export.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        stand_report = client.get(
            f"/api/runs/{run_id}/exports/node-anomaly-stand-report.xlsx",
            params={"node_type": ["OpenCargoDoor", "CloseCargoDoor"], "minimum_quantity": 2},
        )
        assert stand_report.status_code == 200
        assert stand_report.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        stand_report_json = client.get(
            f"/api/runs/{run_id}/exports/node-anomaly-stand-report.json",
            params={"node_type": ["OpenCargoDoor", "CloseCargoDoor"], "minimum_quantity": 2},
        )
        assert stand_report_json.status_code == 200
        assert stand_report_json.json()["filters"]["minimum_quantity"] == 2
        stand_statistics = client.get(
            f"/api/runs/{run_id}/exports/node-anomaly-stand-statistics.xlsx",
            params={"node_type": ["OpenCargoDoor", "CloseCargoDoor"]},
        )
        assert stand_statistics.status_code == 200
        statistics_sheet = load_workbook(
            BytesIO(stand_statistics.content), read_only=True
        ).active
        assert [cell.value for cell in next(statistics_sheet.iter_rows())][:3] == [
            "机位号",
            "错误节点1",
            "重复数1",
        ]
        expected_sample_count = min(
            5,
            sum(
                group["assignment_status"] == "NEEDS_REVIEW"
                and "AMBIGUOUS_MATCH" in group["issue_tags"]
                for group in groups
            ),
        )
        assert expected_sample_count > 0
        sampled = client.post(f"/api/runs/{run_id}/acdm-validation/samples", params={"limit": 5})
        assert sampled.status_code == 200
        assert sampled.json()["added"] == expected_sample_count
        assert len(sampled.json()["selected_codes"]) == expected_sample_count
        selected_codes = set(sampled.json()["selected_codes"])
        repeated_sample = client.post(
            f"/api/runs/{run_id}/acdm-validation/samples", params={"limit": 5}
        )
        assert repeated_sample.status_code == 200
        assert repeated_sample.json()["added"] == 0
        assert set(repeated_sample.json()["selected_codes"]) == selected_codes
        initial_validation = client.get(f"/api/runs/{run_id}/acdm-validation").json()
        assert initial_validation["total_cases"] == expected_sample_count
        assert initial_validation["pending_cases"] == expected_sample_count
        assert all(
            item["sample_status"] == "AWAITING_REVIEW" for item in initial_validation["cases"]
        )

        blocked_acceptance = client.get(f"/api/runs/{run_id}/acceptance").json()
        assert blocked_acceptance["can_publish"] is False
        assert isinstance(blocked_acceptance["regression_cases"], list)
        assert "闭环核验样本尚未全部审核" in blocked_acceptance["blockers"]
        blocked_publish = client.post(
            f"/api/strategies/{dashboard.json()['strategy']['id']}/publish",
            params={"run_id": run_id},
        )
        assert blocked_publish.status_code == 409

        candidate_group = None
        candidate_detail = None
        for group in groups:
            if group["temporary_code"] not in selected_codes:
                continue
            current_detail = client.get(f"/api/groups/{group['id']}").json()
            if current_detail["candidates"]:
                candidate_group = group
                candidate_detail = current_detail
                break
        assert candidate_group and candidate_detail
        candidate = candidate_detail["candidates"][0]
        flight_no = candidate["flight_plan"]["inbound_flight_no"]
        simulated = client.post(
            "/api/references/acdm/simulate",
            json={
                "batch_id": 1,
                "temporary_code": candidate_group["temporary_code"],
                "flight_no": flight_no,
                "aircraft_entry_time": candidate_group["observed_start"],
                "chock_on_time": candidate_group["observed_start"],
                "stand_release_time": candidate_group["observed_end"],
            },
        )
        assert simulated.status_code == 200
        assert simulated.json()["flight_no"] == flight_no
        validation = client.get(f"/api/runs/{dashboard.json()['active_run']['id']}/acdm-validation")
        assert validation.status_code == 200
        assert validation.json()["total_cases"] == expected_sample_count
        assert validation.json()["pending_cases"] == expected_sample_count
        simulated_case = next(
            item
            for item in validation.json()["cases"]
            if item["temporary_code"] == candidate_group["temporary_code"]
        )
        assert simulated_case["sample_status"] == "AWAITING_REVIEW"

        cleared = client.delete(
            "/api/references/acdm/simulate",
            params={"batch_id": 1, "temporary_code": candidate_group["temporary_code"]},
        )
        assert cleared.status_code == 200
        assert cleared.json()["deleted"] == 1

        premature_review = client.post(
            f"/api/groups/{candidate_group['id']}/reviews",
            json={"verdict": "correct", "comment": "清除A-CDM后允许直接提交人工答案"},
        )
        assert premature_review.status_code == 400

        review = client.post(
            f"/api/groups/{missing['id']}/reviews",
            json={"verdict": "unassigned", "comment": "确认作为无航班号临时组"},
        )
        assert review.status_code == 200
        assert review.json()["expected_assignment_status"] == "UNASSIGNED"

        association_groups = client.get(f"/api/runs/{run_id}/associations/groups")
        assert association_groups.status_code == 200
        association_rows = association_groups.json()
        assert association_rows
        assert association_rows[0]["nodes"] == []
        association_detail = client.get(
            f"/api/runs/{run_id}/associations/groups",
            params={"group_id": association_rows[0]["group_id"], "include_nodes": True},
        )
        assert association_detail.status_code == 200
        assert association_detail.json()[0]["node_count"] == len(
            association_detail.json()[0]["nodes"]
        )

        flight_associations = client.get(f"/api/runs/{run_id}/associations/flights")
        assert flight_associations.status_code == 200
        assert all("@2026-" in item["association_key"] for item in flight_associations.json())

        exported = client.get(f"/api/runs/{run_id}/exports/unassigned.json")
        assert exported.status_code == 200
        assert all(item["flight_no"] is None for item in exported.json())

        rerun = client.post(
            "/api/runs",
            json={"batch_id": 1, "strategy_version_id": 1},
        )
        assert rerun.status_code == 200
        rerun_id = rerun.json()["id"]
        acceptance = client.get(f"/api/runs/{rerun_id}/acceptance")
        assert acceptance.status_code == 200
        assert acceptance.json()["regression_count"] == 0
        rerun_validation = client.get(f"/api/runs/{rerun_id}/acdm-validation")
        assert rerun_validation.status_code == 200
        assert rerun_validation.json()["total_cases"] == expected_sample_count


def test_registration_similarity_api_preserves_requested_ranking() -> None:
    with TestClient(app) as client:
        close = client.post(
            "/api/features/registration-similarity",
            json={"observed": "b533", "candidate": "b53b"},
        )
        farther = client.post(
            "/api/features/registration-similarity",
            json={"observed": "b533", "candidate": "b524"},
        )

        assert close.status_code == 200
        assert close.json()["observed_normalized"] == "B533"
        assert close.json()["similarity"] > farther.json()["similarity"]


def test_recovery_queue_and_xian_outbound_policy_are_machine_driven() -> None:
    with TestClient(app) as client:
        queue = client.get("/api/recovery-groups")
        assert queue.status_code == 200
        payload = queue.json()
        assert payload["total"] == len(payload["items"])
        assert payload["statistics"]["unresolved"] == 0
        assert all(item["machine_status"] != "REVIEW_REQUIRED" for item in payload["items"])
        pending = next(
            item for item in payload["items"] if item["machine_status"] == "RECOVERY_PENDING"
        )
        pending_detail = client.get(
            f"/api/recovery-groups/{pending['group_id']}"
        ).json()
        assert pending_detail["attempt_count"] == 1
        assert len(pending_detail["attempts"]) == 1
        assert pending_detail["attempts"][0]["status"] == "NO_RESPONSE"
        assert pending_detail["next_attempt_at"] is not None
        assert payload["statistics"].get("MATCHED_RECOVERED", 0) == 0

        matched = next(item for item in payload["items"] if item["machine_status"] == "MATCHED")
        assert matched["outbox_status"] == "PREVIEWED"

        unassigned = next(
            item for item in payload["items"] if item["machine_status"] == "UNASSIGNED_FINAL"
        )
        preview = client.get(
            f"/api/recovery-groups/{unassigned['group_id']}/payload-preview"
        )
        assert preview.status_code == 200
        assert preview.json()["outbound_status"] == "SUPPRESSED_BY_POLICY"
        assert preview.json()["payload"] is None
        detail = client.get(f"/api/recovery-groups/{unassigned['group_id']}").json()
        if detail["reason_code"] == "RECOVERY_EXHAUSTED":
            assert len(detail["attempts"]) == detail["max_attempts"]

        policy = client.get("/api/recovery-policies/effective")
        assert policy.status_code == 200
        assert policy.json()["airport_code"] == "XIY"
        assert policy.json()["config"]["temporary_group_send_enabled"] is False
        assert policy.json()["temporary_group_send_locked"] is True


def test_recovery_policy_requires_versioned_replay_before_publish() -> None:
    with TestClient(app) as client:
        effective = client.get("/api/recovery-policies/effective").json()
        conflict = client.post(
            "/api/recovery-policies/drafts",
            json={
                "airport_code": "XIY",
                "expected_version": effective["version"] + 1,
                "idempotency_key": "policy-conflict",
                "config": {},
            },
        )
        assert conflict.status_code == 409

        request = {
            "airport_code": "XIY",
            "tenant_code": "default",
            "destination": "xian_bus",
            "expected_version": effective["version"],
            "idempotency_key": "policy-draft-v2",
            "config": {"temporary_group_send_enabled": True, "max_attempts": 4},
        }
        draft = client.post("/api/recovery-policies/drafts", json=request)
        repeated = client.post("/api/recovery-policies/drafts", json=request)
        assert draft.status_code == 200
        assert repeated.json()["id"] == draft.json()["id"]
        assert draft.json()["config"]["temporary_group_send_enabled"] is False
        assert draft.json()["config"]["max_attempts"] == 4

        replay_request = {
            "run_id": client.get("/api/dashboard").json()["active_run"]["id"],
            "idempotency_key": "policy-replay-v2",
        }
        replay = client.post(
            f"/api/recovery-policies/drafts/{draft.json()['id']}/replays",
            json=replay_request,
        )
        replay_again = client.post(
            f"/api/recovery-policies/drafts/{draft.json()['id']}/replays",
            json=replay_request,
        )
        assert replay.status_code == 200
        assert replay.json()["status"] == "SUCCEEDED"
        assert replay_again.json()["id"] == replay.json()["id"]
        assert replay.json()["evidence"]["node_conservation"] is True

        approved = client.post(
            f"/api/recovery-policies/drafts/{draft.json()['id']}/approve",
            json={"replay_task_id": replay.json()["id"]},
        )
        assert approved.json()["status"] == "approved"
        published = client.post(
            f"/api/recovery-policies/drafts/{draft.json()['id']}/publish",
            json={
                "idempotency_key": "publish-v2",
                "expected_version": effective["version"],
            },
        )
        assert published.json()["status"] == "published"
        assert client.get("/api/recovery-policies/effective").json()["id"] == draft.json()["id"]


def test_cluster_review_is_independent_from_flight_match_review() -> None:
    with TestClient(app) as client:
        groups = client.get("/api/groups").json()
        target = next(group for group in groups if group["node_count"] >= 3)
        detail = client.get(f"/api/groups/{target['id']}").json()
        split_node_id = detail["nodes"][1]["id"]

        reviewed = client.post(
            f"/api/groups/{target['id']}/cluster-reviews",
            json={"verdict": "correct", "comment": "边界与节点顺序均正确"},
        )
        assert reviewed.status_code == 200
        assert reviewed.json()["verdict"] == "correct"

        refreshed = client.get(f"/api/groups/{target['id']}").json()
        assert refreshed["cluster_review_status"] == "correct"
        assert refreshed["cluster_reviews"][-1]["comment"] == "边界与节点顺序均正确"
        assert refreshed["review_status"] == target["review_status"]

        split_review = client.post(
            f"/api/groups/{target['id']}/cluster-reviews",
            json={"verdict": "split_required", "split_node_id": split_node_id},
        )
        assert split_review.status_code == 200
        assert split_review.json()["split_node_id"] == split_node_id

        invalid_split = client.post(
            f"/api/groups/{target['id']}/cluster-reviews",
            json={"verdict": "split_required", "split_node_id": detail["nodes"][0]["id"]},
        )
        assert invalid_split.status_code == 400

        anomaly = client.post(
            f"/api/groups/{target['id']}/cluster-reviews",
            json={"verdict": "anomaly", "anomaly_node_ids": [detail["nodes"][0]["id"]]},
        )
        assert anomaly.status_code == 200
        assert anomaly.json()["anomaly_node_ids"] == [detail["nodes"][0]["id"]]


def test_split_replaces_superseded_recovery_records_without_duplicate_nodes() -> None:
    with TestClient(app) as client:
        run_id = client.get("/api/dashboard").json()["active_run"]["id"]
        groups = client.get("/api/groups").json()
        target = next(group for group in groups if group["node_count"] >= 4)
        detail = client.get(f"/api/groups/{target['id']}").json()

        response = client.post(
            f"/api/groups/{target['id']}/split",
            params={"split_node_id": detail["nodes"][2]["id"]},
        )

        assert response.status_code == 200
        assert len(response.json()) == 2
        assert client.get(f"/api/recovery-groups/{target['id']}").status_code == 404
        node_ids = [node["id"] for node in detail["nodes"]]
        placeholders = ",".join("?" for _ in node_ids)
        with sqlite3.connect("test_flight_simulator.db") as connection:
            duplicate = connection.execute(
                f"""
                SELECT disposition.node_id
                FROM recovery_node_dispositions AS disposition
                JOIN flight_groups AS flight_group ON flight_group.id = disposition.group_id
                WHERE flight_group.assignment_status != 'SUPERSEDED'
                  AND flight_group.run_id = ?
                  AND disposition.node_id IN ({placeholders})
                GROUP BY disposition.node_id
                HAVING COUNT(*) > 1
                LIMIT 1
                """,
                [run_id, *node_ids],
            ).fetchone()
        assert duplicate is None


def test_acdm_can_confirm_and_review_flight_without_plan() -> None:
    with TestClient(app) as client:
        dashboard = client.get("/api/dashboard").json()
        groups = client.get("/api/groups").json()
        missing = next(group for group in groups if "MISSING_PLAN" in group["issue_tags"])
        simulated = client.post(
            "/api/references/acdm/simulate",
            json={
                "batch_id": dashboard["batch"]["id"],
                "temporary_code": missing["temporary_code"],
                "flight_no": "DIV123",
                "aircraft_entry_time": missing["observed_start"],
                "chock_on_time": missing["observed_start"],
                "stand_release_time": missing["observed_end"],
            },
        )
        assert simulated.status_code == 200
        rerun = client.post(
            "/api/runs",
            json={
                "batch_id": dashboard["batch"]["id"],
                "strategy_version_id": dashboard["strategy"]["id"],
            },
        ).json()
        rerun_groups = client.get(f"/api/groups?run_id={rerun['id']}").json()
        confirmed = next(
            group for group in rerun_groups if group["temporary_code"] == missing["temporary_code"]
        )
        assert confirmed["assignment_status"] == "MATCHED_REFERENCE_NO_PLAN"

        review = client.post(
            f"/api/groups/{confirmed['id']}/reviews",
            json={
                "verdict": "incorrect",
                "error_type": "flight_match_error",
                "correct_flight_no": "DIV123",
            },
        )
        assert review.status_code == 200
        assert review.json()["expected_flight_no"] == "DIV123"
        validation = client.get(f"/api/runs/{rerun['id']}/acdm-validation").json()
        case = next(item for item in validation["cases"] if item["acdm_flight_no"] == "DIV123")
        assert case["current_strategy_correct"] is True
        assert case["acdm_matches_final"] is True
        assert case["final_status"] == "MATCHED_REFERENCE_NO_PLAN"

        associations = client.get(f"/api/runs/{rerun['id']}/associations/flights").json()
        assert any(item["flight_no"] == "DIV123" for item in associations)


def test_acdm_closed_loop_uses_pre_review_snapshot_and_detects_regression() -> None:
    with TestClient(app) as client:
        dashboard = client.get("/api/dashboard").json()
        groups = client.get("/api/groups").json()
        source_group = None
        source_detail = None
        for group in groups:
            detail = client.get(f"/api/groups/{group['id']}").json()
            if detail["candidates"] and not detail["acdm_reference"] and not detail["reviews"]:
                source_group = group
                source_detail = detail
                break
        assert source_group and source_detail
        source_candidate = source_detail["candidates"][0]["flight_plan"]
        source_flight_no = source_candidate["inbound_flight_no"]
        reference_payload = {
            "batch_id": dashboard["batch"]["id"],
            "temporary_code": source_group["temporary_code"],
            "flight_no": source_flight_no,
            "aircraft_entry_time": source_group["observed_start"],
            "chock_on_time": source_group["observed_start"],
            "stand_release_time": source_group["observed_end"],
        }
        assert (
            client.post("/api/references/acdm/simulate", json=reference_payload).status_code == 200
        )
        initial_run = client.post(
            "/api/runs",
            json={
                "batch_id": dashboard["batch"]["id"],
                "strategy_version_id": dashboard["strategy"]["id"],
            },
        ).json()
        initial_group = next(
            group
            for group in client.get(f"/api/groups?run_id={initial_run['id']}").json()
            if group["temporary_code"] == source_group["temporary_code"]
        )

        final_flight_no = "ZZ999"
        review = client.post(
            f"/api/groups/{initial_group['id']}/reviews",
            json={
                "verdict": "incorrect",
                "error_type": "flight_match_error",
                "correct_flight_no": final_flight_no,
            },
        )
        assert review.status_code == 200
        initial_validation = client.get(f"/api/runs/{initial_run['id']}/acdm-validation").json()
        initial_case = next(
            item
            for item in initial_validation["cases"]
            if item["temporary_code"] == source_group["temporary_code"]
        )
        assert initial_case["current_strategy_correct"] is False
        assert initial_case["current_flight_no"] != final_flight_no
        assert initial_case["acdm_matches_final"] is False
        assert initial_case["is_regression"] is False
        assert initial_validation["review_errors"] == 1
        initial_suggestions = client.get(f"/api/runs/{initial_run['id']}/suggestions").json()
        assert any(item["key"] == "review_acdm_reference_conflicts" for item in initial_suggestions)
        assert not any("reference" in item.get("patch", {}).get("weights", {}) for item in initial_suggestions)

        fixed_payload = {**reference_payload, "flight_no": final_flight_no}
        assert client.post("/api/references/acdm/simulate", json=fixed_payload).status_code == 200
        fixed_run = client.post(
            "/api/runs",
            json={
                "batch_id": dashboard["batch"]["id"],
                "strategy_version_id": dashboard["strategy"]["id"],
            },
        ).json()
        fixed_group = next(
            group
            for group in client.get(f"/api/groups?run_id={fixed_run['id']}").json()
            if group["temporary_code"] == source_group["temporary_code"]
        )
        fixed_validation = client.get(f"/api/runs/{fixed_run['id']}/acdm-validation").json()
        fixed_case = next(
            item
            for item in fixed_validation["cases"]
            if item["temporary_code"] == source_group["temporary_code"]
        )
        assert fixed_case["current_strategy_correct"] is True
        assert fixed_validation["review_errors"] == 0
        assert fixed_validation["regression_count"] == 0
        fixed_acceptance = client.get(f"/api/runs/{fixed_run['id']}/acceptance").json()
        fixed_regression = next(
            item
            for item in fixed_acceptance["regression_cases"]
            if item["temporary_code"] == source_group["temporary_code"]
        )
        assert fixed_regression["expected_result"] == final_flight_no
        assert fixed_regression["passed"] is True
        fixed_suggestions = client.get(f"/api/runs/{fixed_run['id']}/suggestions").json()
        assert not any(
            item["key"] == "review_acdm_reference_conflicts"
            for item in fixed_suggestions
        )

        certified = client.post(
            f"/api/groups/{fixed_group['id']}/reviews",
            json={"verdict": "correct", "comment": "确认修复结果"},
        )
        assert certified.status_code == 200
        assert (
            client.post("/api/references/acdm/simulate", json=reference_payload).status_code == 200
        )
        regressed_run = client.post(
            "/api/runs",
            json={
                "batch_id": dashboard["batch"]["id"],
                "strategy_version_id": dashboard["strategy"]["id"],
            },
        ).json()
        regressed_validation = client.get(f"/api/runs/{regressed_run['id']}/acdm-validation").json()
        regressed_case = next(
            item
            for item in regressed_validation["cases"]
            if item["temporary_code"] == source_group["temporary_code"]
        )
        assert regressed_case["current_strategy_correct"] is False
        assert regressed_case["is_regression"] is True
        assert regressed_validation["regression_count"] == 1
        acceptance = client.get(f"/api/runs/{regressed_run['id']}/acceptance").json()
        assert acceptance["can_publish"] is False
        assert "历史正确案例发生回退" in acceptance["blockers"]
        regressed_result = next(
            item
            for item in acceptance["regression_cases"]
            if item["temporary_code"] == source_group["temporary_code"]
        )
        assert regressed_result["expected_result"] == final_flight_no
        assert regressed_result["passed"] is False
