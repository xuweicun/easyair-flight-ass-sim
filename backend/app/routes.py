from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db import get_db
from app.importers import parse_flight_plans, parse_nodes
from app.node_anomalies import EVENT_LABELS, detect_node_anomalies, summarize_stand_node_issues
from app.models import (
    AppearanceFeature,
    AcdmReferenceFeature,
    Batch,
    ClusterReview,
    FlightGroup,
    FlightPlan,
    GroupNode,
    NodeEvent,
    Review,
    RecoveryDelivery,
    RecoveryPolicy,
    RecoveryReplayTask,
    RecoveryResolution,
    SimulationRun,
    StrategyVersion,
    ValidationSample,
    utcnow,
)
from app.recovery import (
    DEFAULT_RECOVERY_CONFIG,
    effective_policy,
    ensure_run_resolutions,
    evaluate_policy_replay,
    next_policy_version,
    process_recovery_cycle,
    recovery_attempts,
)
from app.registration_similarity import normalize_registration, registration_similarity
from app.schemas import (
    AcceptanceOut,
    AcdmReferenceIn,
    AcdmReferenceOut,
    AcdmValidationSummaryOut,
    AssociationGroupOut,
    AppearanceIn,
    AppearanceOut,
    BatchOut,
    ClusterReviewIn,
    ClusterReviewOut,
    DashboardOut,
    GroupDetailOut,
    GroupListOut,
    FlightAssociationOut,
    ReviewIn,
    ReviewOut,
    RegistrationSimilarityIn,
    RegistrationSimilarityOut,
    RecoveryApproveIn,
    RecoveryPolicyDraftIn,
    RecoveryPublishIn,
    RecoveryReplayIn,
    RunIn,
    RunOut,
    StrategyCreate,
    StrategyOut,
    SuggestionOut,
)
from app.services import (
    association_groups,
    execute_run,
    group_detail,
    latest_run,
    merge_groups,
    node_phase,
    review_counts,
    split_group,
)

router = APIRouter(prefix="/api")


def _policy_payload(policy: RecoveryPolicy) -> dict[str, Any]:
    return {
        "id": policy.id,
        "airport_code": policy.airport_code,
        "tenant_code": policy.tenant_code,
        "destination": policy.destination,
        "version": policy.version,
        "status": policy.status,
        "config": policy.config,
        "temporary_group_send_locked": policy.airport_code == "XIY",
        "created_at": policy.created_at,
        "approved_at": policy.approved_at,
        "published_at": policy.published_at,
    }


def _resolution_payload(
    resolution: RecoveryResolution,
    group: FlightGroup,
    batch: Batch,
    node_count: int | None = None,
) -> dict[str, Any]:
    return {
        "id": resolution.id,
        "group_id": group.id,
        "temporary_code": group.temporary_code,
        "airport_code": batch.airport_code,
        "stand": group.stand,
        "observed_start": group.observed_start,
        "observed_end": group.observed_end,
        "node_count": len(group.nodes) if node_count is None else node_count,
        "group_version": resolution.group_version,
        "member_hash": resolution.member_hash,
        "machine_status": resolution.machine_status,
        "reason_code": resolution.reason_code,
        "attempt_count": resolution.attempt_count,
        "max_attempts": resolution.max_attempts,
        "recovery_deadline": resolution.recovery_deadline,
        "next_attempt_at": resolution.next_attempt_at,
        "last_attempt_at": resolution.last_attempt_at,
        "recovery_request_id": resolution.recovery_request_id,
        "request_window_start": resolution.request_window_start,
        "request_window_end": resolution.request_window_end,
        "response_flight_count": resolution.response_flight_count,
        "candidates": resolution.candidate_summary or [],
        "outbound_policy": resolution.outbound_policy,
        "outbox_status": resolution.outbox_status,
        "config_version": resolution.config_version,
        "strategy_version": resolution.strategy_version,
        "finalized_at": resolution.finalized_at,
        "issue_tags": group.issue_tags or [],
    }


def _review_status(group: FlightGroup) -> str:
    return group.reviews[-1].verdict if group.reviews else "pending"


def _cluster_review_status(group: FlightGroup) -> str:
    return group.cluster_reviews[-1].verdict if group.cluster_reviews else "pending"


def _group_list(group: FlightGroup) -> GroupListOut:
    return GroupListOut(
        id=group.id,
        temporary_code=group.temporary_code,
        stand=group.stand,
        observed_start=group.observed_start,
        observed_end=group.observed_end,
        assignment_status=group.assignment_status,
        assigned_flight_id=group.assigned_flight_id,
        confidence=group.confidence,
        margin=group.margin,
        issue_tags=group.issue_tags or [],
        lineage=group.lineage or {},
        node_count=len(group.nodes),
        review_status=_review_status(group),
        cluster_review_status=_cluster_review_status(group),
    )


async def _appearance_for_group(db: AsyncSession, group: FlightGroup) -> AppearanceFeature | None:
    run = await db.get(SimulationRun, group.run_id)
    if not run:
        return None
    return (
        await db.execute(
            select(AppearanceFeature)
            .where(AppearanceFeature.batch_id == run.batch_id)
            .where(AppearanceFeature.temporary_code == group.temporary_code)
            .order_by(AppearanceFeature.id.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


async def _acdm_reference_for_group(
    db: AsyncSession, group: FlightGroup
) -> AcdmReferenceFeature | None:
    run = await db.get(SimulationRun, group.run_id)
    if not run:
        return None
    return (
        await db.execute(
            select(AcdmReferenceFeature)
            .where(AcdmReferenceFeature.batch_id == run.batch_id)
            .where(AcdmReferenceFeature.temporary_code == group.temporary_code)
            .order_by(AcdmReferenceFeature.id.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


def _flight_label(plan: FlightPlan | None) -> str | None:
    if not plan:
        return None
    return (
        " / ".join(value for value in (plan.inbound_flight_no, plan.outbound_flight_no) if value)
        or None
    )


async def _attributed_flight_numbers(db: AsyncSession, group: FlightGroup | None) -> set[str]:
    if not group:
        return set()
    if group.assigned_flight_id is not None:
        plan = await db.get(FlightPlan, group.assigned_flight_id)
        return {
            value.upper()
            for value in (
                plan.inbound_flight_no if plan else None,
                plan.outbound_flight_no if plan else None,
            )
            if value
        }
    if group.assignment_status == "MATCHED_REFERENCE_NO_PLAN":
        flight_no = (group.lineage or {}).get("acdm_reference", {}).get("flight_no")
        return {flight_no.upper()} if flight_no else set()
    return set()


async def _group_node_ids(db: AsyncSession, group_id: int) -> list[int]:
    return list(
        (
            await db.execute(
                select(GroupNode.node_id)
                .where(GroupNode.group_id == group_id)
                .order_by(GroupNode.order_index)
            )
        ).scalars()
    )


async def _ensure_validation_sample(
    db: AsyncSession, batch_id: int, run_id: int, group: FlightGroup
) -> ValidationSample:
    sample = (
        await db.execute(
            select(ValidationSample)
            .where(ValidationSample.batch_id == batch_id)
            .where(ValidationSample.temporary_code == group.temporary_code)
        )
    ).scalar_one_or_none()
    if sample:
        return sample
    sample = ValidationSample(
        batch_id=batch_id,
        temporary_code=group.temporary_code,
        source_run_id=run_id,
        node_ids=await _group_node_ids(db, group.id),
    )
    db.add(sample)
    await db.flush()
    return sample


async def _group_for_sample(
    db: AsyncSession,
    run_id: int,
    sample: ValidationSample,
    current_by_code: dict[str, FlightGroup],
) -> FlightGroup | None:
    exact = current_by_code.get(sample.temporary_code)
    if exact or not sample.node_ids:
        return exact
    expected = set(sample.node_ids)
    candidate_ids = list(
        (
            await db.execute(
                select(GroupNode.group_id)
                .join(FlightGroup, FlightGroup.id == GroupNode.group_id)
                .where(FlightGroup.run_id == run_id)
                .where(GroupNode.node_id.in_(expected))
                .group_by(GroupNode.group_id)
                .having(func.count(GroupNode.node_id) == len(expected))
            )
        ).scalars()
    )
    for group_id in candidate_ids:
        if set(await _group_node_ids(db, group_id)) == expected:
            return await db.get(FlightGroup, group_id)
    return None


async def _validation_sample_for_group(
    db: AsyncSession, group: FlightGroup
) -> ValidationSample | None:
    run = await db.get(SimulationRun, group.run_id)
    if not run:
        return None
    samples = list(
        (
            await db.execute(
                select(ValidationSample).where(ValidationSample.batch_id == run.batch_id)
            )
        ).scalars()
    )
    exact = next(
        (sample for sample in samples if sample.temporary_code == group.temporary_code),
        None,
    )
    if exact:
        return exact
    node_ids = set(await _group_node_ids(db, group.id))
    return next(
        (sample for sample in samples if node_ids and set(sample.node_ids or []) == node_ids),
        None,
    )


@router.get("/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@router.get("/dashboard", response_model=DashboardOut)
async def dashboard(db: AsyncSession = Depends(get_db)) -> DashboardOut:
    batch = (await db.execute(select(Batch).order_by(Batch.id.desc()).limit(1))).scalar_one()
    run = await latest_run(db, batch.id)
    if not run:
        raise HTTPException(404, "该批次尚未运行策略")
    strategy = await db.get(StrategyVersion, run.strategy_version_id)
    counts = await review_counts(db, run.id)
    return DashboardOut(
        batch=batch,
        active_run=run,
        strategy=strategy,
        issue_counts=run.metrics.get("issue_counts", {}),
        validation={
            "required_reviews": counts["required"],
            "completed_reviews": counts["completed"],
            "incorrect_reviews": counts["incorrect"],
            "node_conservation": run.metrics.get("node_conservation", False),
        },
    )


@router.get("/batches", response_model=list[BatchOut])
async def batches(db: AsyncSession = Depends(get_db)) -> list[Batch]:
    return list((await db.execute(select(Batch).order_by(Batch.id.desc()))).scalars())


@router.post("/batches/import", response_model=RunOut)
async def import_batch(
    name: str = Form(...),
    airport_code: str = Form("XIY"),
    plan_file: UploadFile = File(...),
    node_file: UploadFile = File(...),
    manual_file: UploadFile | None = File(None),
    acdm_file: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
) -> SimulationRun:
    plan_content = await plan_file.read()
    node_content = await node_file.read()
    manual_content = await manual_file.read() if manual_file else None
    acdm_content = await acdm_file.read() if acdm_file else None
    plans = parse_flight_plans(plan_content)
    nodes = parse_nodes(node_content, "algorithm_node")
    manual_nodes = parse_nodes(manual_content, "manual_report") if manual_content else []
    acdm_nodes = parse_nodes(acdm_content, "acdm_reference") if acdm_content else []
    if not plans or not nodes:
        raise HTTPException(400, "航班计划或节点文件没有可识别的数据")

    batch = Batch(
        name=name,
        airport_code=airport_code,
        status="ready",
        source_files={
            "plan": {
                "name": plan_file.filename,
                "sha256": hashlib.sha256(plan_content).hexdigest(),
            },
            "node": {
                "name": node_file.filename,
                "sha256": hashlib.sha256(node_content).hexdigest(),
            },
            "acdm_manual": {"name": manual_file.filename} if manual_file else None,
            "acdm": {"name": acdm_file.filename} if acdm_file else None,
        },
        stats={
            "plan_groups": len(plans),
            "nodes": len(nodes),
            "acdm_nodes": len(manual_nodes) + len(acdm_nodes),
        },
    )
    db.add(batch)
    await db.flush()
    db.add_all([FlightPlan(batch_id=batch.id, **item) for item in plans])
    db.add_all(
        [NodeEvent(batch_id=batch.id, **item) for item in [*nodes, *manual_nodes, *acdm_nodes]]
    )
    await db.commit()
    strategy = (
        await db.execute(
            select(StrategyVersion)
            .where(StrategyVersion.status == "published")
            .order_by(StrategyVersion.id.desc())
            .limit(1)
        )
    ).scalar_one()
    return await execute_run(db, batch, strategy)


@router.post("/references/acdm/import")
async def import_acdm(
    batch_id: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    batch = await db.get(Batch, batch_id)
    if not batch:
        raise HTTPException(404, "批次不存在")
    content = await file.read()
    rows = parse_nodes(content, "acdm_reference")
    db.add_all([NodeEvent(batch_id=batch.id, **item) for item in rows])
    await db.commit()
    return {"imported": len(rows), "requires_rerun": True}


@router.get("/groups", response_model=list[GroupListOut])
async def groups(
    run_id: int | None = None,
    status: str | None = None,
    issue: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> list[GroupListOut]:
    if run_id is None:
        run = await latest_run(db)
        if not run:
            return []
        run_id = run.id
    stmt = (
        select(FlightGroup)
        .where(FlightGroup.run_id == run_id)
        .where(FlightGroup.assignment_status != "SUPERSEDED")
        .options(
            selectinload(FlightGroup.nodes),
            selectinload(FlightGroup.reviews),
            selectinload(FlightGroup.cluster_reviews),
        )
        .order_by(FlightGroup.observed_start)
    )
    if status:
        stmt = stmt.where(FlightGroup.assignment_status == status)
    result = list((await db.execute(stmt)).scalars())
    if issue:
        result = [group for group in result if issue in (group.issue_tags or [])]
    return [_group_list(group) for group in result]


@router.get("/groups/{group_id}", response_model=GroupDetailOut)
async def get_group(group_id: int, db: AsyncSession = Depends(get_db)) -> GroupDetailOut:
    group = await group_detail(db, group_id)
    if not group:
        raise HTTPException(404, "分组不存在")
    appearance = await _appearance_for_group(db, group)
    acdm_reference = await _acdm_reference_for_group(db, group)
    assigned_plan = next(
        (
            candidate.flight_plan
            for candidate in group.candidates
            if candidate.flight_plan_id == group.assigned_flight_id
        ),
        None,
    )
    if not assigned_plan and group.assigned_flight_id:
        assigned_plan = await db.get(FlightPlan, group.assigned_flight_id)
    related_segments = await _related_flight_segments(db, group, assigned_plan)
    base = _group_list(group).model_dump()
    return GroupDetailOut(
        **base,
        nodes=[
            {
                **{
                    column: getattr(item.node, column)
                    for column in (
                        "id",
                        "source_type",
                        "event_type",
                        "event_time",
                        "stand",
                        "reported_flight_no",
                        "safeguard_code",
                        "is_anomaly",
                    )
                },
                "phase": node_phase(item.node.event_type),
                "attributed_flight_no": (
                    assigned_plan.inbound_flight_no
                    if assigned_plan and node_phase(item.node.event_type) == "ARRIVAL"
                    else assigned_plan.outbound_flight_no
                    if assigned_plan
                    else None
                ),
            }
            for item in sorted(group.nodes, key=lambda value: value.order_index)
        ],
        candidates=sorted(group.candidates, key=lambda value: value.rank),
        reviews=sorted(group.reviews, key=lambda value: value.id),
        cluster_reviews=sorted(group.cluster_reviews, key=lambda value: value.id),
        related_segments=related_segments,
        appearance=appearance,
        acdm_reference=acdm_reference,
    )


async def _node_anomaly_items(db: AsyncSession, run_id: int) -> list[dict[str, Any]]:
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(404, "策略运行不存在")
    groups = list(
        (
            await db.execute(
                select(FlightGroup)
                .where(FlightGroup.run_id == run_id)
                .where(FlightGroup.assignment_status != "SUPERSEDED")
                .options(selectinload(FlightGroup.nodes).selectinload(GroupNode.node))
                .order_by(FlightGroup.observed_start)
            )
        ).scalars()
    )
    return detect_node_anomalies(groups)


@router.get("/runs/{run_id}/node-anomalies")
async def node_anomalies(
    run_id: int,
    problem_code: str | None = None,
    stand: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    all_items = await _node_anomaly_items(db, run_id)
    items = [
        item
        for item in all_items
        if (not problem_code or item["problem_code"] == problem_code)
        and (not stand or stand.upper() in item["stand"].upper())
    ]
    type_counts: dict[str, int] = {}
    for item in all_items:
        type_counts[item["problem_type"]] = type_counts.get(item["problem_type"], 0) + 1
    return {
        "run_id": run_id,
        "repeat_window_minutes": 5,
        "statistics": {
            "total": len(all_items),
            "affected_stands": len({item["stand"] for item in all_items}),
            "rapid_repeat": sum(
                item["problem_code"] == "RAPID_REPEAT" for item in all_items
            ),
            "guide_car_only": sum(
                item["problem_code"] == "GUIDE_CAR_ONLY" for item in all_items
            ),
            "by_type": type_counts,
        },
        "items": items,
    }


@router.get("/runs/{run_id}/exports/node-anomalies.json")
async def export_node_anomalies_json(
    run_id: int, db: AsyncSession = Depends(get_db)
) -> JSONResponse:
    items = await _node_anomaly_items(db, run_id)
    return JSONResponse(
        json.loads(json.dumps(items, default=lambda value: value.isoformat()))
    )


@router.get("/runs/{run_id}/exports/node-anomalies.xlsx")
async def export_node_anomalies_excel(
    run_id: int, db: AsyncSession = Depends(get_db)
) -> StreamingResponse:
    items = await _node_anomaly_items(db, run_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "节点异常"
    sheet.append(
        [
            "机位号",
            "问题类型",
            "问题原因",
            "临时保障组",
            "异常开始时间",
            "异常结束时间",
            "异常节点数",
            "保障组节点数",
            "节点类型",
        ]
    )
    for item in items:
        sheet.append(
            [
                item["stand"],
                item["problem_type"],
                item["reason"],
                item["temporary_code"],
                item["window_start"].isoformat(sep=" ", timespec="seconds"),
                item["window_end"].isoformat(sep=" ", timespec="seconds"),
                item["affected_node_count"],
                item["group_node_count"],
                ", ".join(item["event_types"]),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {"A": 12, "B": 20, "C": 58, "D": 34, "E": 22, "F": 22, "G": 14, "H": 14, "I": 28}.items():
        sheet.column_dimensions[column].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="node-anomalies-run-{run_id}.xlsx"'
        },
    )


@router.get("/runs/{run_id}/exports/node-anomaly-stand-report.xlsx")
async def export_node_anomaly_stand_report(
    run_id: int,
    node_type: list[str] | None = Query(None),
    minimum_quantity: int = Query(1, ge=1),
    problem_code: str | None = None,
    stand: str | None = None,
    query: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    items = await _node_anomaly_items(db, run_id)
    filtered_items = _filter_node_anomaly_items(items, problem_code, stand, query)
    rows = summarize_stand_node_issues(
        filtered_items,
        node_types=set(node_type or []),
        minimum_quantity=minimum_quantity,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "机位问题报告"
    sheet.append(
        [
            "机位",
            "问题节点",
            "问题航班数量",
            "异常节点数量",
            "问题航班/保障组",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row["stand"],
                row["event_label"],
                row["problem_flight_count"],
                row["anomaly_node_count"],
                "、".join(row["temporary_codes"]),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {"A": 12, "B": 22, "C": 18, "D": 18, "E": 72}.items():
        sheet.column_dimensions[column].width = width
    criteria = workbook.create_sheet("筛选条件")
    criteria.append(["筛选项", "筛选值"])
    criteria.append(["节点类型", "、".join(EVENT_LABELS.get(value, value) for value in (node_type or [])) or "全部异常节点"])
    criteria.append(["最少异常节点数量", minimum_quantity])
    criteria.append(["问题类型", problem_code or "全部问题类型"])
    criteria.append(["机位", stand or "全部机位"])
    criteria.append(["搜索条件", query or "无"])
    criteria.column_dimensions["A"].width = 22
    criteria.column_dimensions["B"].width = 56
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="node-anomaly-stand-report-run-{run_id}.xlsx"'
            )
        },
    )


@router.get("/runs/{run_id}/exports/node-anomaly-stand-report.json")
async def export_node_anomaly_stand_report_json(
    run_id: int,
    node_type: list[str] | None = Query(None),
    minimum_quantity: int = Query(1, ge=1),
    problem_code: str | None = None,
    stand: str | None = None,
    query: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    items = _filter_node_anomaly_items(
        await _node_anomaly_items(db, run_id), problem_code, stand, query
    )
    rows = summarize_stand_node_issues(
        items,
        node_types=set(node_type or []),
        minimum_quantity=minimum_quantity,
    )
    return JSONResponse(
        {
            "run_id": run_id,
            "filters": {
                "node_types": node_type or [],
                "minimum_quantity": minimum_quantity,
                "problem_code": problem_code,
                "stand": stand,
                "query": query,
            },
            "items": rows,
        }
    )


@router.get("/runs/{run_id}/exports/node-anomaly-stand-statistics.xlsx")
async def export_node_anomaly_stand_statistics(
    run_id: int,
    node_type: list[str] | None = Query(None),
    minimum_quantity: int = Query(1, ge=1),
    problem_code: str | None = None,
    stand: str | None = None,
    query: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    items = _filter_node_anomaly_items(
        await _node_anomaly_items(db, run_id), problem_code, stand, query
    )
    rows = summarize_stand_node_issues(
        items,
        node_types=set(node_type or []),
        minimum_quantity=minimum_quantity,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "机位节点统计"
    rows_by_stand: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        rows_by_stand.setdefault(row["stand"], []).append(row)
    max_node_types = max(
        1, max((len(items) for items in rows_by_stand.values()), default=0)
    )
    headers = ["机位号"]
    for index in range(1, max_node_types + 1):
        headers.extend([f"错误节点{index}", f"重复数{index}"])
    sheet.append(headers)
    for stand, stand_rows in rows_by_stand.items():
        values: list[Any] = [stand]
        for row in stand_rows:
            values.extend([row["event_label"], row["anomaly_node_count"]])
        sheet.append(values)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 14
    for index in range(max_node_types):
        node_column = 2 + index * 2
        count_column = node_column + 1
        sheet.column_dimensions[get_column_letter(node_column)].width = 24
        sheet.column_dimensions[get_column_letter(count_column)].width = 14
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="node-anomaly-stand-statistics-run-{run_id}.xlsx"'
            )
        },
    )


def _filter_node_anomaly_items(
    items: list[dict[str, Any]],
    problem_code: str | None,
    stand: str | None,
    query: str | None,
) -> list[dict[str, Any]]:
    query_text = (query or "").strip().upper()
    return [
        item
        for item in items
        if (not problem_code or item["problem_code"] == problem_code)
        and (not stand or item["stand"] == stand)
        and (
            not query_text
            or query_text in item["stand"].upper()
            or query_text in item["temporary_code"].upper()
        )
    ]


async def _related_flight_segments(
    db: AsyncSession,
    current_group: FlightGroup,
    current_plan: FlightPlan | None,
) -> list[dict[str, Any]]:
    if not current_plan:
        return []
    flight_numbers = {
        value.upper()
        for value in (current_plan.inbound_flight_no, current_plan.outbound_flight_no)
        if value
    }
    if not flight_numbers:
        return []
    service_day_start = (
        current_plan.plan_start.replace(hour=0, minute=0, second=0, microsecond=0)
        if current_plan.plan_start
        else current_group.observed_start.replace(hour=0, minute=0, second=0, microsecond=0)
    )
    service_day_end = service_day_start + timedelta(days=1)
    rows = (
        await db.execute(
            select(FlightGroup, FlightPlan)
            .join(FlightPlan, FlightPlan.id == FlightGroup.assigned_flight_id)
            .where(FlightGroup.run_id == current_group.run_id)
            .where(FlightGroup.assignment_status != "SUPERSEDED")
            .where(FlightPlan.plan_start >= service_day_start)
            .where(FlightPlan.plan_start < service_day_end)
            .where(
                or_(
                    FlightPlan.inbound_flight_no.in_(flight_numbers),
                    FlightPlan.outbound_flight_no.in_(flight_numbers),
                )
            )
            .options(selectinload(FlightGroup.nodes).selectinload(GroupNode.node))
            .order_by(FlightGroup.observed_start, FlightGroup.id)
        )
    ).all()
    segments: list[dict[str, Any]] = []
    for group, plan in rows:
        phases = [node_phase(item.node.event_type) for item in group.nodes]
        for flight_no, phase, node_count in (
            (plan.inbound_flight_no, "ARRIVAL", phases.count("ARRIVAL")),
            (
                plan.outbound_flight_no,
                "OUTBOUND",
                sum(value in {"TURNAROUND", "DEPARTURE"} for value in phases),
            ),
        ):
            if flight_no and node_count:
                segments.append(
                    {
                        "group_id": group.id,
                        "temporary_code": group.temporary_code,
                        "flight_no": flight_no,
                        "phase": phase,
                        "stand": group.stand,
                        "aircraft_no": plan.aircraft_no,
                        "node_count": node_count,
                        "current_group": group.id == current_group.id,
                    }
                )
    aggregated: dict[tuple[str, str | None, str, str], dict[str, Any]] = {}
    for segment in segments:
        key = (
            segment["stand"],
            segment["aircraft_no"],
            segment["flight_no"],
            segment["phase"],
        )
        existing = aggregated.get(key)
        if not existing:
            aggregated[key] = segment
            continue
        existing["node_count"] += segment["node_count"]
        if segment["current_group"]:
            existing.update(
                group_id=segment["group_id"],
                temporary_code=segment["temporary_code"],
                current_group=True,
            )
    return list(aggregated.values())


@router.post("/references/acdm/simulate", response_model=AcdmReferenceOut)
async def simulate_acdm_reference(
    payload: AcdmReferenceIn, db: AsyncSession = Depends(get_db)
) -> AcdmReferenceFeature:
    if not await db.get(Batch, payload.batch_id):
        raise HTTPException(404, "批次不存在")
    feature = AcdmReferenceFeature(
        batch_id=payload.batch_id,
        temporary_code=payload.temporary_code,
        flight_no=payload.flight_no.strip().upper(),
        node_payload={
            "aircraft_entry_time": payload.aircraft_entry_time.isoformat(),
            "chock_on_time": (payload.chock_on_time.isoformat() if payload.chock_on_time else None),
            "stand_release_time": (
                payload.stand_release_time.isoformat() if payload.stand_release_time else None
            ),
        },
        source_type="acdm_simulation",
    )
    db.add(feature)
    run = await latest_run(db, payload.batch_id)
    if run:
        group = (
            await db.execute(
                select(FlightGroup)
                .where(FlightGroup.run_id == run.id)
                .where(FlightGroup.temporary_code == payload.temporary_code)
            )
        ).scalar_one_or_none()
        if group:
            await _ensure_validation_sample(db, payload.batch_id, run.id, group)
    await db.commit()
    await db.refresh(feature)
    return feature


@router.delete("/references/acdm/simulate")
async def clear_acdm_reference(
    batch_id: int, temporary_code: str, db: AsyncSession = Depends(get_db)
) -> dict[str, int]:
    result = await db.execute(
        delete(AcdmReferenceFeature)
        .where(AcdmReferenceFeature.batch_id == batch_id)
        .where(AcdmReferenceFeature.temporary_code == temporary_code)
    )
    await db.commit()
    return {"deleted": result.rowcount or 0}


@router.post("/runs/{run_id}/acdm-validation/samples")
async def select_acdm_validation_samples(
    run_id: int, limit: int = 5, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(404, "运行不存在")
    target = max(1, min(limit, 50))
    groups = list(
        (
            await db.execute(
                select(FlightGroup)
                .where(FlightGroup.run_id == run_id)
                .where(FlightGroup.assignment_status != "SUPERSEDED")
            )
        ).scalars()
    )
    groups_by_code = {group.temporary_code: group for group in groups}
    samples = list(
        (
            await db.execute(
                select(ValidationSample)
                .where(ValidationSample.batch_id == run.batch_id)
                .order_by(ValidationSample.selected_at, ValidationSample.id)
            )
        ).scalars()
    )
    selected_codes = {sample.temporary_code for sample in samples}
    feature_codes = list(
        (
            await db.execute(
                select(AcdmReferenceFeature.temporary_code)
                .where(AcdmReferenceFeature.batch_id == run.batch_id)
                .distinct()
            )
        ).scalars()
    )
    added = 0
    for code in feature_codes:
        group = groups_by_code.get(code)
        if group and code not in selected_codes:
            await _ensure_validation_sample(db, run.batch_id, run_id, group)
            selected_codes.add(code)
            added += 1
    candidates = sorted(
        (
            group
            for group in groups
            if group.temporary_code not in selected_codes
            and group.assignment_status == "NEEDS_REVIEW"
            and "AMBIGUOUS_MATCH" in (group.issue_tags or [])
        ),
        key=lambda group: (group.margin, group.observed_start, group.id),
    )
    for group in candidates:
        if len(selected_codes) >= target:
            break
        await _ensure_validation_sample(db, run.batch_id, run_id, group)
        selected_codes.add(group.temporary_code)
        added += 1
    await db.commit()
    selected = list(
        (
            await db.execute(
                select(ValidationSample)
                .where(ValidationSample.batch_id == run.batch_id)
                .order_by(ValidationSample.selected_at, ValidationSample.id)
            )
        ).scalars()
    )
    return {
        "selected_codes": [sample.temporary_code for sample in selected],
        "added": added,
    }


@router.get(
    "/runs/{run_id}/acdm-validation",
    response_model=AcdmValidationSummaryOut,
)
async def acdm_validation_summary(
    run_id: int, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(404, "运行不存在")

    feature_rows = list(
        (
            await db.execute(
                select(AcdmReferenceFeature)
                .where(AcdmReferenceFeature.batch_id == run.batch_id)
                .order_by(AcdmReferenceFeature.id)
            )
        ).scalars()
    )
    latest_features: dict[str, AcdmReferenceFeature] = {}
    for feature in feature_rows:
        latest_features[feature.temporary_code] = feature

    sample_rows = list(
        (
            await db.execute(
                select(ValidationSample)
                .where(ValidationSample.batch_id == run.batch_id)
                .order_by(ValidationSample.selected_at, ValidationSample.id)
            )
        ).scalars()
    )
    samples_by_code = {sample.temporary_code: sample for sample in sample_rows}
    case_codes = [sample.temporary_code for sample in sample_rows]
    for feature in feature_rows:
        if feature.temporary_code not in samples_by_code:
            samples_by_code[feature.temporary_code] = ValidationSample(
                batch_id=run.batch_id,
                temporary_code=feature.temporary_code,
                source_run_id=run_id,
                node_ids=[],
            )
            case_codes.append(feature.temporary_code)

    current_groups = list(
        (
            await db.execute(
                select(FlightGroup)
                .where(FlightGroup.run_id == run_id)
                .where(FlightGroup.assignment_status != "SUPERSEDED")
            )
        ).scalars()
    )
    current_by_code = {group.temporary_code: group for group in current_groups}
    review_rows = (
        await db.execute(
            select(Review, FlightGroup)
            .join(FlightGroup, FlightGroup.id == Review.group_id)
            .join(SimulationRun, SimulationRun.id == FlightGroup.run_id)
            .where(SimulationRun.batch_id == run.batch_id)
            .order_by(Review.id)
        )
    ).all()
    latest_reviews: dict[str, tuple[Review, FlightGroup]] = {}
    for review, review_group in review_rows:
        if review.expected_assignment_status in {None, "NEEDS_REVIEW", "SUPERSEDED"}:
            continue
        latest_reviews[review_group.temporary_code] = (review, review_group)

    cases = []
    for temporary_code in case_codes:
        sample = samples_by_code[temporary_code]
        feature = latest_features.get(temporary_code)
        group = await _group_for_sample(db, run_id, sample, current_by_code)
        review_record = latest_reviews.get(temporary_code) or (
            latest_reviews.get(group.temporary_code) if group else None
        )
        review, review_group = review_record if review_record else (None, None)
        baseline_group = (
            (
                await db.execute(
                    select(FlightGroup)
                    .join(SimulationRun, SimulationRun.id == FlightGroup.run_id)
                    .where(SimulationRun.batch_id == run.batch_id)
                    .where(SimulationRun.created_at < feature.created_at)
                    .where(FlightGroup.temporary_code == temporary_code)
                    .where(FlightGroup.assignment_status != "SUPERSEDED")
                    .order_by(SimulationRun.created_at.desc())
                    .limit(1)
                )
            ).scalar_one_or_none()
            if feature
            else None
        )
        final_plan = (
            await db.get(FlightPlan, review.expected_flight_id)
            if review and review.expected_flight_id
            else None
        )
        baseline_plan = (
            await db.get(FlightPlan, baseline_group.assigned_flight_id)
            if baseline_group and baseline_group.assigned_flight_id
            else None
        )
        current_plan = (
            await db.get(FlightPlan, group.assigned_flight_id)
            if group and group.assigned_flight_id
            else None
        )
        final_numbers = {
            value.upper()
            for value in (
                final_plan.inbound_flight_no if final_plan else None,
                final_plan.outbound_flight_no if final_plan else None,
                review.expected_flight_no if review else None,
            )
            if value
        }
        current_numbers = await _attributed_flight_numbers(db, group)
        baseline_numbers = await _attributed_flight_numbers(db, baseline_group)
        acdm_matches_final = (
            feature.flight_no.upper() in final_numbers if review and feature else None
        )
        review_comparison = (
            (review_group.lineage or {}).get("latest_review_comparison", {}) if review_group else {}
        )
        same_run_review = bool(group and review_group and group.id == review_group.id)
        snapshot_correct = review_comparison.get("strategy_correct")
        if not review or not group:
            current_correct = None
        elif same_run_review and isinstance(snapshot_correct, bool):
            current_correct = snapshot_correct
        elif review.expected_flight_id is not None:
            current_correct = group.assigned_flight_id == review.expected_flight_id
        elif review.expected_flight_no:
            current_correct = review.expected_flight_no.upper() in current_numbers
        else:
            current_correct = group.assignment_status == review.expected_assignment_status
        if not review or not baseline_group:
            baseline_correct = None
        elif review.expected_flight_id is not None:
            baseline_correct = baseline_group.assigned_flight_id == review.expected_flight_id
        elif review.expected_flight_no:
            baseline_correct = review.expected_flight_no.upper() in baseline_numbers
        else:
            baseline_correct = baseline_group.assignment_status == review.expected_assignment_status
        resolved_by_acdm = baseline_correct is False and current_correct is True if review else None
        source_strategy_correct = review_comparison.get("strategy_correct")
        is_regression = bool(
            review_group
            and review_group.run_id != run_id
            and source_strategy_correct is True
            and current_correct is False
        )
        if not review:
            sample_status = "AWAITING_REVIEW"
        elif is_regression:
            sample_status = "REGRESSION"
        elif current_correct is False or not group:
            sample_status = "NEEDS_STRATEGY_FIX"
        else:
            sample_status = "VALIDATED"
        current_flight_id = group.assigned_flight_id if group else None
        current_flight_no = (
            (group.lineage or {}).get("acdm_reference", {}).get("flight_no")
            if group and not current_plan
            else _flight_label(current_plan)
        )
        if same_run_review:
            snapshot_flight_id = review_comparison.get("strategy_flight_id")
            snapshot_plan = (
                await db.get(FlightPlan, snapshot_flight_id)
                if isinstance(snapshot_flight_id, int)
                else None
            )
            current_flight_id = snapshot_flight_id if isinstance(snapshot_flight_id, int) else None
            current_flight_no = review_comparison.get("strategy_flight_no") or _flight_label(
                snapshot_plan
            )
        cases.append(
            {
                "temporary_code": group.temporary_code if group else temporary_code,
                "group_id": group.id if group else None,
                "acdm_flight_no": feature.flight_no if feature else None,
                "sample_status": sample_status,
                "baseline_flight_id": (
                    baseline_group.assigned_flight_id if baseline_group else None
                ),
                "baseline_flight_no": _flight_label(baseline_plan),
                "baseline_status": (baseline_group.assignment_status if baseline_group else None),
                "current_flight_id": current_flight_id,
                "current_flight_no": current_flight_no,
                "final_flight_id": review.expected_flight_id if review else None,
                "final_flight_no": (
                    review.expected_flight_no if review else _flight_label(final_plan)
                ),
                "final_status": review.expected_assignment_status if review else None,
                "review_verdict": review.verdict if review else None,
                "acdm_matches_final": acdm_matches_final,
                "baseline_strategy_correct": baseline_correct,
                "current_strategy_correct": current_correct,
                "resolved_by_acdm": resolved_by_acdm,
                "is_regression": is_regression,
            }
        )

    reviewed = [case for case in cases if case["review_verdict"] is not None]
    pending = [
        case for case in cases if case["sample_status"] in {"AWAITING_ACDM", "AWAITING_REVIEW"}
    ]
    return {
        "run_id": run_id,
        "total_cases": len(cases),
        "reviewed_cases": len(reviewed),
        "pending_cases": len(pending),
        "review_errors": sum(case["current_strategy_correct"] is False for case in reviewed),
        "acdm_conflicts": sum(case["acdm_matches_final"] is False for case in reviewed),
        "baseline_error_count": sum(
            case["baseline_strategy_correct"] is False for case in reviewed
        ),
        "resolved_by_acdm_count": sum(case["resolved_by_acdm"] is True for case in reviewed),
        "regression_count": sum(case["is_regression"] is True for case in reviewed),
        "cases": cases,
    }


@router.get("/runs/{run_id}/associations/groups", response_model=list[AssociationGroupOut])
async def associations_by_group(
    run_id: int,
    overrun_only: bool = False,
    overrun_minutes: int = 20,
    group_id: int | None = None,
    include_nodes: bool = False,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    if not await db.get(SimulationRun, run_id):
        raise HTTPException(404, "运行不存在")
    rows = await association_groups(db, run_id)
    if group_id is not None:
        rows = [row for row in rows if row["group_id"] == group_id]
    if overrun_only:
        rows = [row for row in rows if row["overrun_minutes"] > overrun_minutes]
    if not include_nodes:
        rows = [{**row, "nodes": []} for row in rows]
    return rows


@router.get("/runs/{run_id}/associations/flights", response_model=list[FlightAssociationOut])
async def associations_by_flight(
    run_id: int,
    query: str | None = None,
    overrun_only: bool = False,
    overrun_minutes: int = 20,
    association_key: str | None = None,
    include_nodes: bool = False,
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    rows = await association_groups(db, run_id)
    by_flight: dict[str, list[dict]] = {}
    for row in rows:
        if row["assignment_status"] == "MATCHED_REFERENCE_NO_PLAN" and row["inbound_flight_no"]:
            flight_no = row["inbound_flight_no"]
            key = f"{flight_no}@{row['service_date']}"
            by_flight.setdefault(key, []).append(row)
            continue
        for flight_no, phases in (
            (row["inbound_flight_no"], {"ARRIVAL"}),
            (row["outbound_flight_no"], {"TURNAROUND", "DEPARTURE"}),
        ):
            if not flight_no:
                continue
            piece = {**row, "nodes": [node for node in row["nodes"] if node["phase"] in phases]}
            if piece["nodes"]:
                key = f"{flight_no}@{row['service_date']}"
                by_flight.setdefault(key, []).append(piece)
    output = []
    normalized_query = (query or "").strip().upper()
    normalized_key = (association_key or "").strip().upper()
    for key, pieces in sorted(by_flight.items()):
        flight_no, service_date = key.rsplit("@", 1)
        if normalized_key and key.upper() != normalized_key:
            continue
        if normalized_query and normalized_query not in flight_no.upper():
            continue
        max_overrun = max(piece["overrun_minutes"] for piece in pieces)
        if overrun_only and max_overrun <= overrun_minutes:
            continue
        aircraft = sorted({piece["aircraft_no"] for piece in pieces if piece["aircraft_no"]})
        output_pieces = pieces if include_nodes else [{**piece, "nodes": []} for piece in pieces]
        output.append(
            {
                "association_key": key,
                "flight_no": flight_no,
                "service_date": service_date,
                "groups": output_pieces,
                "stands": sorted({piece["stand"] for piece in pieces}),
                "aircraft": aircraft,
                "has_aircraft_change": len(aircraft) > 1,
                "max_overrun_minutes": max_overrun,
            }
        )
    return output


@router.post("/groups/{group_id}/reviews", response_model=ReviewOut)
async def create_review(
    group_id: int, payload: ReviewIn, db: AsyncSession = Depends(get_db)
) -> Review:
    group = await db.get(FlightGroup, group_id)
    if not group:
        raise HTTPException(404, "分组不存在")
    if payload.verdict == "incorrect" and not payload.error_type:
        raise HTTPException(400, "指出错误时必须选择错误类型")
    acdm_reference = await _acdm_reference_for_group(db, group)
    strategy_flight_id = group.assigned_flight_id
    strategy_status = group.assignment_status
    if payload.verdict == "correct" and strategy_status not in {
        "MATCHED",
        "MATCHED_REFERENCE",
        "MATCHED_REFERENCE_NO_PLAN",
        "MATCHED_MANUAL",
    }:
        raise HTTPException(
            400,
            "当前策略尚未形成明确航班；请选择最终航班号或保留无航班号",
        )
    strategy_plan = (
        await db.get(FlightPlan, strategy_flight_id) if strategy_flight_id is not None else None
    )
    strategy_numbers = await _attributed_flight_numbers(db, group)
    strategy_flight_no = _flight_label(strategy_plan)
    if not strategy_flight_no and strategy_status == "MATCHED_REFERENCE_NO_PLAN":
        strategy_flight_no = (group.lineage or {}).get("acdm_reference", {}).get("flight_no")
    correct_flight_no = (
        payload.correct_flight_no.strip().upper() if payload.correct_flight_no else None
    )
    if payload.correct_flight_id:
        plan = await db.get(FlightPlan, payload.correct_flight_id)
        if not plan:
            raise HTTPException(400, "纠正航班不存在")
        plan_numbers = [
            value.upper() for value in (plan.inbound_flight_no, plan.outbound_flight_no) if value
        ]
        if correct_flight_no and correct_flight_no not in plan_numbers:
            raise HTTPException(400, "最终航班号不属于所选航班计划")
        correct_flight_no = correct_flight_no or (plan_numbers[0] if plan_numbers else None)
        group.assigned_flight_id = plan.id
        group.assignment_status = "MATCHED_MANUAL"
    elif correct_flight_no:
        group.assigned_flight_id = None
        group.assignment_status = "MATCHED_REFERENCE_NO_PLAN"
        group.lineage = {
            **(group.lineage or {}),
            "acdm_reference": {
                **(group.lineage or {}).get("acdm_reference", {}),
                "state": "confirmed_plan_missing",
                "flight_no": correct_flight_no,
            },
        }
    elif payload.verdict == "unassigned":
        group.assigned_flight_id = None
        group.assignment_status = "UNASSIGNED"
    elif payload.verdict == "data_error":
        group.assigned_flight_id = None
        group.assignment_status = "DATA_ERROR"
    expected_flight_id = group.assigned_flight_id
    expected_status = group.assignment_status
    expected_plan = (
        await db.get(FlightPlan, expected_flight_id) if expected_flight_id is not None else None
    )
    expected_flight_no = correct_flight_no
    if not expected_flight_no and expected_plan:
        expected_plan_numbers = [
            value.upper()
            for value in (expected_plan.inbound_flight_no, expected_plan.outbound_flight_no)
            if value
        ]
        if acdm_reference and acdm_reference.flight_no.upper() in expected_plan_numbers:
            expected_flight_no = acdm_reference.flight_no.upper()
        else:
            expected_flight_no = expected_plan_numbers[0] if expected_plan_numbers else None
    if not expected_flight_no and payload.verdict == "correct" and acdm_reference:
        expected_flight_no = acdm_reference.flight_no.upper()
    acdm_matches_final = bool(
        acdm_reference
        and expected_flight_no
        and acdm_reference.flight_no.upper() == expected_flight_no.upper()
    )
    group.lineage = {
        **(group.lineage or {}),
        "latest_review_comparison": {
            "strategy_flight_id": strategy_flight_id,
            "strategy_flight_no": strategy_flight_no,
            "strategy_status": strategy_status,
            "final_flight_id": expected_flight_id,
            "final_flight_no": expected_flight_no,
            "final_status": expected_status,
            "strategy_correct": (
                expected_flight_no.upper() in strategy_numbers
                if expected_flight_no
                else strategy_flight_id == expected_flight_id and strategy_status == expected_status
            ),
            "acdm_flight_no": acdm_reference.flight_no if acdm_reference else None,
            "acdm_matches_final": acdm_matches_final if acdm_reference else None,
        },
    }
    review = Review(
        group_id=group_id,
        expected_flight_id=expected_flight_id,
        expected_flight_no=expected_flight_no,
        expected_assignment_status=expected_status,
        **payload.model_dump(exclude={"correct_flight_no"}),
        correct_flight_no=correct_flight_no,
    )
    db.add(review)
    await db.commit()
    await db.refresh(review)
    return review


@router.post("/groups/{group_id}/cluster-reviews", response_model=ClusterReviewOut)
async def create_cluster_review(
    group_id: int, payload: ClusterReviewIn, db: AsyncSession = Depends(get_db)
) -> ClusterReview:
    group = await group_detail(db, group_id)
    if not group:
        raise HTTPException(404, "分组不存在")
    ordered_node_ids = [
        item.node_id for item in sorted(group.nodes, key=lambda value: value.order_index)
    ]
    if payload.verdict == "split_required":
        if payload.split_node_id not in ordered_node_ids[1:]:
            raise HTTPException(400, "拆分节点必须属于当前组且不能是第一个节点")
    if payload.verdict == "merge_required":
        merge_ids = sorted(set(payload.merge_group_ids))
        if group_id not in merge_ids or len(merge_ids) < 2:
            raise HTTPException(400, "合并审核必须包含当前组和至少一个相邻组")
        merge_groups = list(
            (
                await db.execute(
                    select(FlightGroup).where(FlightGroup.id.in_(merge_ids))
                )
            ).scalars()
        )
        if len(merge_groups) != len(merge_ids):
            raise HTTPException(400, "合并审核包含不存在的分组")
        if any(
            item.run_id != group.run_id
            or item.stand != group.stand
            or item.assignment_status == "SUPERSEDED"
            for item in merge_groups
        ):
            raise HTTPException(400, "只能审核同一运行、同一机位的有效分组")
    if payload.verdict == "anomaly":
        anomaly_ids = sorted(set(payload.anomaly_node_ids))
        if not anomaly_ids or any(node_id not in ordered_node_ids for node_id in anomaly_ids):
            raise HTTPException(400, "异常节点必须来自当前分组")
    review = ClusterReview(group_id=group_id, **payload.model_dump())
    db.add(review)
    await db.commit()
    await db.refresh(review)
    return review


@router.post("/groups/{group_id}/split", response_model=list[GroupListOut])
async def split(
    group_id: int, split_node_id: int, db: AsyncSession = Depends(get_db)
) -> list[GroupListOut]:
    try:
        created = await split_group(db, group_id, split_node_id)
    except ValueError as error:
        raise HTTPException(400, str(error)) from error
    output = []
    for group in created:
        detail = await group_detail(db, group.id)
        output.append(_group_list(detail))
    return output


@router.post("/groups/merge", response_model=GroupListOut)
async def merge(group_ids: list[int], db: AsyncSession = Depends(get_db)) -> GroupListOut:
    try:
        group = await merge_groups(db, group_ids)
    except ValueError as error:
        raise HTTPException(400, str(error)) from error
    detail = await group_detail(db, group.id)
    return _group_list(detail)


@router.post("/features/appearance", response_model=AppearanceOut)
async def create_appearance(
    payload: AppearanceIn, db: AsyncSession = Depends(get_db)
) -> AppearanceFeature:
    batch = await db.get(Batch, payload.batch_id)
    if not batch:
        raise HTTPException(404, "批次不存在")
    feature = AppearanceFeature(**payload.model_dump())
    db.add(feature)
    await db.commit()
    await db.refresh(feature)
    return feature


@router.post("/features/registration-similarity", response_model=RegistrationSimilarityOut)
async def compare_aircraft_registration(
    payload: RegistrationSimilarityIn,
) -> RegistrationSimilarityOut:
    return RegistrationSimilarityOut(
        observed_normalized=normalize_registration(payload.observed),
        candidate_normalized=normalize_registration(payload.candidate),
        similarity=registration_similarity(payload.observed, payload.candidate),
    )


@router.get("/recovery-groups")
async def recovery_groups(
    run_id: int | None = None,
    status: str | None = None,
    stand: str | None = None,
    reason: str | None = None,
    outbound_status: str | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
    config_version: int | None = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    try:
        run, batch, policy, _ = await ensure_run_resolutions(db, run_id)
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    filters = [
        FlightGroup.run_id == run.id,
        FlightGroup.assignment_status != "SUPERSEDED",
    ]
    if status:
        filters.append(RecoveryResolution.machine_status == status)
    if stand:
        filters.append(FlightGroup.stand == stand)
    if reason:
        filters.append(RecoveryResolution.reason_code == reason)
    if outbound_status:
        filters.append(RecoveryResolution.outbox_status == outbound_status)
    if config_version is not None:
        filters.append(RecoveryResolution.config_version == config_version)
    if start_time:
        filters.append(FlightGroup.observed_end >= datetime.fromisoformat(start_time))
    if end_time:
        filters.append(FlightGroup.observed_start <= datetime.fromisoformat(end_time))
    total = int(await db.scalar(
        select(func.count(RecoveryResolution.id))
        .join(FlightGroup, FlightGroup.id == RecoveryResolution.group_id)
        .where(*filters)
    ) or 0)
    rows = (
        await db.execute(
            select(RecoveryResolution, FlightGroup, func.count(GroupNode.id))
            .join(FlightGroup, FlightGroup.id == RecoveryResolution.group_id)
            .outerjoin(GroupNode, GroupNode.group_id == FlightGroup.id)
            .where(*filters)
            .group_by(RecoveryResolution.id, FlightGroup.id)
            .order_by(FlightGroup.observed_start)
            .offset(offset)
            .limit(limit)
        )
    ).all()
    items = [
        _resolution_payload(resolution, group, batch, int(node_count))
        for resolution, group, node_count in rows
    ]
    now = utcnow()
    status_counts = {
        key: int(value)
        for key, value in (
            await db.execute(
                select(RecoveryResolution.machine_status, func.count(RecoveryResolution.id))
                .join(FlightGroup, FlightGroup.id == RecoveryResolution.group_id)
                .where(FlightGroup.run_id == run.id)
                .group_by(RecoveryResolution.machine_status)
            )
        ).all()
    }
    overdue = int(await db.scalar(
        select(func.count(RecoveryResolution.id))
        .join(FlightGroup, FlightGroup.id == RecoveryResolution.group_id)
        .where(
            FlightGroup.run_id == run.id,
            RecoveryResolution.machine_status == "RECOVERY_PENDING",
            RecoveryResolution.recovery_deadline < now,
        )
    ) or 0)
    suppressed = int(await db.scalar(
        select(func.count(RecoveryResolution.id))
        .join(FlightGroup, FlightGroup.id == RecoveryResolution.group_id)
        .where(
            FlightGroup.run_id == run.id,
            RecoveryResolution.outbox_status == "SUPPRESSED_BY_POLICY",
        )
    ) or 0)
    outbox_dead = int(await db.scalar(
        select(func.count(RecoveryDelivery.id))
        .join(FlightGroup, FlightGroup.id == RecoveryDelivery.group_id)
        .where(FlightGroup.run_id == run.id, RecoveryDelivery.outbox_status == "DEAD")
    ) or 0)
    outbox_overdue = int(await db.scalar(
        select(func.count(RecoveryDelivery.id))
        .join(FlightGroup, FlightGroup.id == RecoveryDelivery.group_id)
        .where(
            FlightGroup.run_id == run.id,
            RecoveryDelivery.outbox_status == "PENDING",
            RecoveryDelivery.created_at
            < now - timedelta(seconds=int(policy.config["outbox_max_wait_seconds"])),
        )
    ) or 0)
    terminal_count = sum(
        value for key, value in status_counts.items() if key != "RECOVERY_PENDING"
    )
    delivery_count = int(await db.scalar(
        select(func.count(RecoveryDelivery.id))
        .join(FlightGroup, FlightGroup.id == RecoveryDelivery.group_id)
        .where(FlightGroup.run_id == run.id)
    ) or 0)
    missing_delivery = max(0, terminal_count - delivery_count)
    return {
        "run_id": run.id,
        "as_of": now,
        "total": total,
        "offset": offset,
        "limit": limit,
        "statistics": {
            **status_counts,
            "timeout": overdue,
            "unresolved": overdue + outbox_dead + outbox_overdue + missing_delivery,
            "outbound_suppressed": suppressed,
            "outbox_dead": outbox_dead,
            "outbox_overdue": outbox_overdue,
            "missing_delivery": missing_delivery,
        },
        "items": items,
    }


@router.get("/recovery-groups/{group_id}/payload-preview")
async def recovery_group_payload_preview(
    group_id: int, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    group = await group_detail(db, group_id)
    resolution = (
        await db.execute(
            select(RecoveryResolution).where(RecoveryResolution.group_id == group_id)
        )
    ).scalar_one_or_none()
    if not group or not resolution:
        raise HTTPException(404, "恢复保障组不存在")
    if resolution.outbox_status == "SUPPRESSED_BY_POLICY":
        return {
            "group_id": group.id,
            "outbound_status": "SUPPRESSED_BY_POLICY",
            "reason": "XIY_BUS_REJECTS_TEMPORARY_GROUP",
            "payload": None,
            "protocol_validation": {
                "valid": False,
                "contract": "NOT_APPLICABLE",
                "message": "西安总线不接受缺少正式航班号的临时保障组",
            },
        }
    return {
        "group_id": group.id,
        "outbound_status": resolution.outbox_status,
        "reason": None,
        "payload": {
            "temporary_group_id": group.temporary_code,
            "flight_plan_id": group.assigned_flight_id,
            "stand": group.stand,
            "nodes": [
                {
                    "id": item.node.id,
                    "event_type": item.node.event_type,
                    "event_time": item.node.event_time,
                }
                for item in sorted(group.nodes, key=lambda row: row.order_index)
            ],
        },
        "protocol_validation": {"valid": group.assigned_flight_id is not None},
    }


@router.get("/recovery-groups/{group_id}")
async def recovery_group_detail(
    group_id: int, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    group = await group_detail(db, group_id)
    if not group:
        raise HTTPException(404, "保障组不存在")
    try:
        _, batch, _, _ = await ensure_run_resolutions(db, group.run_id)
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    resolution = (
        await db.execute(
            select(RecoveryResolution).where(RecoveryResolution.group_id == group_id)
        )
    ).scalar_one_or_none()
    if not resolution:
        raise HTTPException(404, "保障组恢复记录不存在或已被新分组替代")
    attempts = await recovery_attempts(db, resolution.id)
    payload = _resolution_payload(resolution, group, batch)
    payload.update(
        {
            "nodes": [
                {
                    "id": item.node.id,
                    "event_type": item.node.event_type,
                    "event_time": item.node.event_time,
                    "source_type": item.node.source_type,
                    "is_anomaly": item.node.is_anomaly,
                }
                for item in sorted(group.nodes, key=lambda row: row.order_index)
            ],
            "cluster_boundary": group.lineage.get("stand_occupancy", {}),
            "first_evaluation": {
                "assignment_status": group.assignment_status,
                "assigned_flight_id": group.assigned_flight_id,
                "confidence": group.confidence,
                "margin": group.margin,
            },
            "attempts": [
                {
                    "attempt_no": item.attempt_no,
                    "request_id": item.request_id,
                    "request_window_start": item.request_window_start,
                    "request_window_end": item.request_window_end,
                    "response_flight_count": item.response_flight_count,
                    "status": item.status,
                    "candidates_before": item.candidates_before,
                    "candidates_after": item.candidates_after,
                    "created_at": item.created_at,
                }
                for item in attempts
            ],
            "status_timeline": resolution.status_timeline or [],
            "raw_audit": group.lineage or {},
        }
    )
    return payload


@router.get("/recovery-policies/effective")
async def get_effective_recovery_policy(
    airport_code: str = "XIY", db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    policy = await effective_policy(db, airport_code)
    await db.commit()
    return _policy_payload(policy)


@router.post("/recovery-policies/drafts")
async def create_recovery_policy_draft(
    payload: RecoveryPolicyDraftIn, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    airport_code = payload.airport_code.upper()
    tenant_code = payload.tenant_code
    destination = payload.destination
    idempotency_key = payload.idempotency_key.strip()
    existing = (
        await db.execute(
            select(RecoveryPolicy).where(RecoveryPolicy.idempotency_key == idempotency_key)
        )
    ).scalar_one_or_none()
    if existing:
        return _policy_payload(existing)
    current = await effective_policy(db, airport_code)
    expected_version = payload.expected_version
    if expected_version != current.version:
        raise HTTPException(
            409,
            detail={"message": "配置版本冲突", "current_version": current.version},
        )
    config = DEFAULT_RECOVERY_CONFIG | payload.config.model_dump(exclude_none=True)
    if airport_code == "XIY":
        config["temporary_group_send_enabled"] = False
        config["destination_capability"] = "UNSUPPORTED"
    config["data_error_recovery_enabled"] = False
    config["recovery_exhausted_disposition"] = "UNASSIGNED_FINAL"
    draft = RecoveryPolicy(
        airport_code=airport_code,
        tenant_code=tenant_code,
        destination=destination,
        version=await next_policy_version(db, airport_code, tenant_code, destination),
        status="draft",
        config=config,
        idempotency_key=idempotency_key,
    )
    db.add(draft)
    await db.commit()
    await db.refresh(draft)
    return _policy_payload(draft)


@router.post("/recovery-policies/drafts/{policy_id}/replays")
async def replay_recovery_policy(
    policy_id: int, payload: RecoveryReplayIn, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    policy = await db.get(RecoveryPolicy, policy_id)
    if not policy or policy.status != "draft":
        raise HTTPException(404, "策略草稿不存在")
    idempotency_key = payload.idempotency_key.strip()
    task = (
        await db.execute(
            select(RecoveryReplayTask).where(
                RecoveryReplayTask.idempotency_key == idempotency_key
            )
        )
    ).scalar_one_or_none()
    if not task:
        run_id = payload.run_id
        counts = await review_counts(db, run_id)
        config = DEFAULT_RECOVERY_CONFIG | dict(policy.config or {})
        try:
            evidence = await evaluate_policy_replay(
                db, run_id, config, counts["regressions"]
            )
        except LookupError as exc:
            raise HTTPException(404, str(exc)) from exc
        task = RecoveryReplayTask(
            policy_id=policy.id,
            idempotency_key=idempotency_key,
            status="SUCCEEDED",
            progress=1,
            evidence=evidence,
            completed_at=utcnow(),
        )
        db.add(task)
        await db.commit()
        await db.refresh(task)
    return {
        "id": task.id,
        "policy_id": task.policy_id,
        "status": task.status,
        "progress": task.progress,
        "evidence": task.evidence,
        "error": task.error,
        "created_at": task.created_at,
        "completed_at": task.completed_at,
    }


@router.get("/replay-tasks/{task_id}")
async def get_replay_task(
    task_id: int, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    task = await db.get(RecoveryReplayTask, task_id)
    if not task:
        raise HTTPException(404, "回放任务不存在")
    return {
        "id": task.id,
        "policy_id": task.policy_id,
        "status": task.status,
        "progress": task.progress,
        "evidence": task.evidence,
        "error": task.error,
        "created_at": task.created_at,
        "completed_at": task.completed_at,
    }


@router.post("/recovery-policies/drafts/{policy_id}/approve")
async def approve_recovery_policy(
    policy_id: int, payload: RecoveryApproveIn, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    policy = await db.get(RecoveryPolicy, policy_id)
    if not policy or policy.status != "draft":
        raise HTTPException(409, "只有草稿可以审批")
    task_id = payload.replay_task_id
    task = await db.get(RecoveryReplayTask, task_id)
    if not task or task.policy_id != policy.id or task.status != "SUCCEEDED":
        raise HTTPException(409, "必须先完成该草稿的回放")
    expected_digest = hashlib.sha256(
        json.dumps(
            DEFAULT_RECOVERY_CONFIG | dict(policy.config or {}),
            sort_keys=True,
            ensure_ascii=False,
        ).encode()
    ).hexdigest()
    if not task.evidence.get("gate_passed"):
        raise HTTPException(409, "回放门禁未通过")
    if task.evidence.get("config_digest") != expected_digest:
        raise HTTPException(409, "回放证据与当前草稿参数不一致")
    policy.status = "approved"
    policy.approved_at = utcnow()
    await db.commit()
    return _policy_payload(policy)


@router.post("/recovery-policies/drafts/{policy_id}/publish")
async def publish_recovery_policy(
    policy_id: int, payload: RecoveryPublishIn, db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    policy = await db.get(RecoveryPolicy, policy_id)
    if policy and policy.status == "published" and policy.publish_idempotency_key == payload.idempotency_key:
        return _policy_payload(policy)
    if not policy or policy.status != "approved":
        raise HTTPException(409, "只有已审批策略可以发布")
    idempotency_key = payload.idempotency_key.strip()
    expected_version = payload.expected_version
    current = await effective_policy(
        db, policy.airport_code, policy.tenant_code, policy.destination
    )
    if expected_version != current.version:
        raise HTTPException(
            409,
            detail={"message": "配置版本冲突", "current_version": current.version},
        )
    prior = list(
        (
            await db.execute(
                select(RecoveryPolicy).where(
                    RecoveryPolicy.airport_code == policy.airport_code,
                    RecoveryPolicy.tenant_code == policy.tenant_code,
                    RecoveryPolicy.destination == policy.destination,
                    RecoveryPolicy.status == "published",
                )
            )
        ).scalars()
    )
    for item in prior:
        item.status = "superseded"
    policy.status = "published"
    policy.published_at = utcnow()
    policy.publish_idempotency_key = idempotency_key
    latest = await latest_run(db)
    if latest:
        await process_recovery_cycle(db, latest.id, commit=False)
    await db.commit()
    return _policy_payload(policy)


@router.get("/strategies", response_model=list[StrategyOut])
async def strategies(db: AsyncSession = Depends(get_db)) -> list[StrategyVersion]:
    return list(
        (await db.execute(select(StrategyVersion).order_by(StrategyVersion.id.desc()))).scalars()
    )


@router.post("/strategies", response_model=StrategyOut)
async def create_strategy(
    payload: StrategyCreate, db: AsyncSession = Depends(get_db)
) -> StrategyVersion:
    strategy = StrategyVersion(**payload.model_dump(), status="draft")
    db.add(strategy)
    await db.commit()
    await db.refresh(strategy)
    return strategy


@router.post("/runs", response_model=RunOut)
async def create_run(payload: RunIn, db: AsyncSession = Depends(get_db)) -> SimulationRun:
    batch = await db.get(Batch, payload.batch_id)
    strategy = await db.get(StrategyVersion, payload.strategy_version_id)
    if not batch or not strategy:
        raise HTTPException(404, "批次或策略不存在")
    return await execute_run(db, batch, strategy)


@router.get("/runs/{run_id}/suggestions", response_model=list[SuggestionOut])
async def suggestions(run_id: int, db: AsyncSession = Depends(get_db)) -> list[SuggestionOut]:
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(404, "运行不存在")
    issues = run.metrics.get("issue_counts", {})
    output = []
    if issues.get("AMBIGUOUS_MATCH"):
        output.append(
            SuggestionOut(
                key="strengthen_semantics",
                title="提高节点阶段语义权重",
                evidence=f"存在 {issues['AMBIGUOUS_MATCH']} 个多候选组，结束节点需要更贴近上一航班结束。",
                patch={"weights": {"node_semantics": 25}},
                affected_groups=issues["AMBIGUOUS_MATCH"],
            )
        )
    if issues.get("INCOMPLETE_SEQUENCE"):
        output.append(
            SuggestionOut(
                key="expand_idle_gap",
                title="检查聚类空闲阈值",
                evidence=f"{issues['INCOMPLETE_SEQUENCE']} 个分组缺少开始或结束边界。",
                patch={"idle_gap_minutes": 150},
                affected_groups=issues["INCOMPLETE_SEQUENCE"],
            )
        )
    if issues.get("LONG_WINDOW") or issues.get("INVALID_YEAR"):
        count = issues.get("LONG_WINDOW", 0) + issues.get("INVALID_YEAR", 0)
        output.append(
            SuggestionOut(
                key="reject_long_window",
                title="收紧异常计划时间窗",
                evidence=f"{count} 个候选受超长或异常年份计划影响，应保持硬拒绝。",
                patch={"hard_reject_plan_hours": 18},
                affected_groups=count,
            )
        )
    if issues.get("MISSING_PLAN"):
        output.append(
            SuggestionOut(
                key="retain_temporary_group",
                title="保留无航班号临时组",
                evidence=f"{issues['MISSING_PLAN']} 个分组没有可信计划，不能丢弃节点。",
                patch={"unmatched_output": "temporary_group"},
                affected_groups=issues["MISSING_PLAN"],
            )
        )
    validation = await acdm_validation_summary(run_id, db)
    reference_conflicts = [
        case
        for case in validation["cases"]
        if case["review_verdict"] is not None and case["acdm_matches_final"] is False
    ]
    if reference_conflicts:
        output.append(
            SuggestionOut(
                key="review_acdm_reference_conflicts",
                title="先复核A-CDM参考冲突",
                evidence=(
                    f"{len(reference_conflicts)} 个案例的A-CDM航班号与最终人工答案不一致，"
                    "应作为辅助证据进入人工复核，不能改变候选基础分。"
                ),
                patch={},
                affected_groups=len(reference_conflicts),
            )
        )
    return output


@router.get("/runs/{run_id}/acceptance", response_model=AcceptanceOut)
async def acceptance(run_id: int, db: AsyncSession = Depends(get_db)) -> AcceptanceOut:
    run = await db.get(SimulationRun, run_id)
    if not run:
        raise HTTPException(404, "运行不存在")
    counts = await review_counts(db, run_id)
    blockers = []
    if counts["completed"] < counts["required"]:
        blockers.append("必审问题尚未全部核验")
    if counts["incorrect"]:
        blockers.append("仍有人工指出的错误")
    if counts["regressions"]:
        blockers.append("历史正确案例发生回退")
    if not run.metrics.get("node_conservation"):
        blockers.append("节点数量不守恒")
    validation = await acdm_validation_summary(run_id, db)
    if validation["total_cases"] == 0:
        blockers.append("尚未建立闭环核验样本")
    elif validation["pending_cases"]:
        blockers.append("闭环核验样本尚未全部审核")
    if validation["review_errors"]:
        blockers.append("A-CDM闭环仍有策略错误")
    if validation["acdm_conflicts"]:
        blockers.append("A-CDM航班号与最终人工答案冲突")
    if validation["regression_count"] and "历史正确案例发生回退" not in blockers:
        blockers.append("历史正确案例发生回退")
    return AcceptanceOut(
        run_id=run_id,
        required_reviews=counts["required"],
        completed_reviews=counts["completed"],
        incorrect_reviews=counts["incorrect"],
        regression_count=counts["regressions"],
        node_conservation=bool(run.metrics.get("node_conservation")),
        can_publish=not blockers,
        blockers=blockers,
        regression_cases=counts["regression_cases"],
    )


@router.post("/strategies/{strategy_id}/publish", response_model=StrategyOut)
async def publish_strategy(
    strategy_id: int, run_id: int, db: AsyncSession = Depends(get_db)
) -> StrategyVersion:
    strategy = await db.get(StrategyVersion, strategy_id)
    run = await db.get(SimulationRun, run_id)
    if not strategy or not run or run.strategy_version_id != strategy.id:
        raise HTTPException(400, "策略和验收运行不匹配")
    gate = await acceptance(run.id, db)
    if not gate.can_publish:
        raise HTTPException(409, "验收门禁尚未通过")
    strategy.status = "published"
    strategy.published_at = utcnow()
    await db.commit()
    await db.refresh(strategy)
    return strategy


async def _unassigned_payloads(db: AsyncSession, run_id: int) -> list[dict[str, Any]]:
    groups = list(
        (
            await db.execute(
                select(FlightGroup)
                .where(FlightGroup.run_id == run_id)
                .where(FlightGroup.assignment_status.in_(("UNASSIGNED", "UNASSIGNED_FINAL")))
                .options(selectinload(FlightGroup.nodes).selectinload(GroupNode.node))
                .order_by(FlightGroup.observed_start)
            )
        ).scalars()
    )
    return [
        {
            "temporary_group_id": group.temporary_code,
            "airport_code": (
                await db.get(Batch, (await db.get(SimulationRun, run_id)).batch_id)
            ).airport_code,
            "stand": group.stand,
            "observed_start": group.observed_start.isoformat(),
            "observed_end": group.observed_end.isoformat(),
            "assignment_status": group.assignment_status,
            "flight_no": None,
            "safeguard_code": None,
            "cluster_confidence": group.confidence,
            "nodes": [
                {
                    "id": item.node.id,
                    "event_type": item.node.event_type,
                    "event_time": item.node.event_time.isoformat()
                    if item.node.event_time
                    else None,
                    "source_type": item.node.source_type,
                }
                for item in sorted(group.nodes, key=lambda value: value.order_index)
            ],
        }
        for group in groups
    ]


@router.get("/runs/{run_id}/exports/unassigned.json")
async def export_unassigned_json(run_id: int, db: AsyncSession = Depends(get_db)) -> JSONResponse:
    return JSONResponse(await _unassigned_payloads(db, run_id))


@router.get("/runs/{run_id}/exports/unassigned.xlsx")
async def export_unassigned_excel(
    run_id: int, db: AsyncSession = Depends(get_db)
) -> StreamingResponse:
    payloads = await _unassigned_payloads(db, run_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "无航班号临时组"
    sheet.append(
        [
            "临时组编号",
            "机场",
            "机位",
            "开始时间",
            "结束时间",
            "节点数",
            "状态",
            "航班号",
            "保障编码",
        ]
    )
    for item in payloads:
        sheet.append(
            [
                item["temporary_group_id"],
                item["airport_code"],
                item["stand"],
                item["observed_start"],
                item["observed_end"],
                len(item["nodes"]),
                item["assignment_status"],
                None,
                None,
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="unassigned-run-{run_id}.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
