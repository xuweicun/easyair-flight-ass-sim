from __future__ import annotations

import json
import calendar
import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import load_workbook


ALIASES: dict[str, tuple[str, ...]] = {
    "safeguard_code": ("safeguardcode", "safeguard_code", "保障编码", "保障编号"),
    "stand": ("stand", "standcode", "stand_code", "机位", "机位号"),
    "flight_no": ("flightno", "flight_no", "航班号", "进港航班", "出港航班"),
    "inbound_flight_no": ("inboundflightno", "in_flight_no", "进港航班号", "进港航班"),
    "outbound_flight_no": ("outboundflightno", "out_flight_no", "出港航班号", "出港航班"),
    "plan_start": (
        "planstart",
        "plan_start",
        "planstarttime",
        "计划开始时间",
        "开始占用时间",
    ),
    "plan_end": (
        "planend",
        "plan_end",
        "planendtime",
        "计划结束时间",
        "结束占用时间",
    ),
    "real_start": ("realstart", "real_start", "actualstart", "实际开始时间"),
    "real_end": ("realend", "real_end", "actualend", "实际结束时间"),
    "event_type": (
        "eventtype",
        "event_type",
        "nodetype",
        "node_type",
        "节点类型",
        "节点名称",
        "事件类型",
    ),
    "event_time": (
        "eventtime",
        "event_time",
        "nodetime",
        "node_time",
        "上报时间",
        "节点时间",
        "事件时间",
    ),
    "source_row_id": ("id", "nodeid", "node_id", "msgid", "消息id", "记录id"),
    "airline": ("airline", "airlinecode", "航司", "所属航司"),
    "aircraft_type": (
        "aircrafttype",
        "aircraft_type",
        "aircraft_model",
        "model",
        "机型",
    ),
    "aircraft_no": (
        "aircraftno",
        "aircraft_no",
        "aircraft_num",
        "tailno",
        "机号",
        "飞机号",
    ),
    "in_out": ("inout", "in_out", "进出港", "进出港标识"),
    "message_content": ("messagecontent", "message_content", "消息内容"),
}


def _key(value: Any) -> str:
    return "".join(str(value or "").strip().lower().replace("-", "_").split())


KNOWN_YEAR_CORRECTIONS = {2076: 2026}


def _correct_datetime_year(value: datetime) -> datetime:
    target_year = KNOWN_YEAR_CORRECTIONS.get(value.year)
    if target_year is None:
        return value
    last_day = calendar.monthrange(target_year, value.month)[1]
    return value.replace(year=target_year, day=min(value.day, last_day))


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return _correct_datetime_year(value).isoformat()
    if isinstance(value, dict):
        return {key: _json_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith(("2076-", "2076/")) or re.fullmatch(r"2076\d{10}", stripped):
            return "2026" + value[value.find("2076") + 4 :]
    return value


def _read_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return []
    output = []
    for row_number, values in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        raw = {headers[index]: _json_value(value) for index, value in enumerate(values)}
        raw["__row_number__"] = row_number
        output.append(raw)
    return output


def _alias_map(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {_key(name): value for name, value in row.items()}
    mapped: dict[str, Any] = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            if _key(alias) in normalized and normalized[_key(alias)] not in (None, ""):
                mapped[canonical] = normalized[_key(alias)]
                break
    return mapped


def _text(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value).strip()


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return _correct_datetime_year(value).replace(tzinfo=None)
    if value is None or value == "":
        return None
    text = str(value).strip()
    for fmt in (
        "%Y%m%d%H%M%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return _correct_datetime_year(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        return _correct_datetime_year(datetime.fromisoformat(text)).replace(tzinfo=None)
    except ValueError:
        return None


def parse_flight_plans(content: bytes) -> list[dict[str, Any]]:
    return normalize_flight_plan_rows(_read_rows(content))


def normalize_flight_plan_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    for raw in rows:
        mapped = _alias_map(raw)
        safeguard = _text(mapped.get("safeguard_code")) or f"ROW-{raw['__row_number__']}"
        stand = _text(mapped.get("stand")) or ""
        grouped[(safeguard, stand)].append((mapped, raw))

    drafts: list[dict[str, Any]] = []
    for (safeguard, stand), records in grouped.items():
        drafts.append(_build_flight_plan(safeguard, stand, records))

    by_occupancy: dict[tuple[str, str, datetime, datetime], list[dict[str, Any]]] = defaultdict(list)
    passthrough: list[dict[str, Any]] = []
    for draft in drafts:
        if all(
            draft.get(key) is not None
            for key in ("aircraft_no", "stand", "plan_start", "plan_end")
        ):
            by_occupancy[
                (
                    draft["aircraft_no"],
                    draft["stand"],
                    draft["plan_start"],
                    draft["plan_end"],
                )
            ].append(draft)
        else:
            passthrough.append(draft)

    plans = list(passthrough)
    for occupancy_key, occupancy_drafts in by_occupancy.items():
        if _is_complementary_occupancy_pair(occupancy_drafts):
            aircraft_no, stand, start, _ = occupancy_key
            records = [
                (_alias_map(raw), raw)
                for draft in occupancy_drafts
                for raw in draft["raw_payload"]["rows"]
            ]
            merged_key = f"OCC-{aircraft_no}-{start:%Y%m%d%H%M%S}"
            plans.append(
                _build_flight_plan(
                    merged_key,
                    stand,
                    records,
                    normalization_method="physical_occupancy",
                )
            )
        else:
            plans.extend(occupancy_drafts)
    return plans


def _build_flight_plan(
    group_key: str,
    stand: str,
    records: list[tuple[dict[str, Any], dict[str, Any]]],
    normalization_method: str = "source_group",
) -> dict[str, Any]:
        starts: list[datetime] = []
        ends: list[datetime] = []
        flight_numbers: list[str] = []
        directions: set[str] = set()
        internal_ids: list[str] = []
        inbound: str | None = None
        outbound: str | None = None
        airline = aircraft_type = aircraft_no = None
        raw_payloads = []
        for mapped, raw in records:
            start = _datetime(mapped.get("real_start")) or _datetime(mapped.get("plan_start"))
            end = _datetime(mapped.get("real_end")) or _datetime(mapped.get("plan_end"))
            if start:
                starts.append(start)
            if end:
                ends.append(end)
            inbound = inbound or _text(mapped.get("inbound_flight_no"))
            outbound = outbound or _text(mapped.get("outbound_flight_no"))
            flight_no = _text(mapped.get("flight_no"))
            if flight_no and flight_no not in flight_numbers:
                flight_numbers.append(flight_no)
            direction = (_text(mapped.get("in_out")) or "").upper()
            if flight_no and direction in {"A", "ARR", "ARRIVAL", "进港"}:
                directions.add("A")
                inbound = inbound or flight_no
            elif flight_no and direction in {"D", "DEP", "DEPARTURE", "出港"}:
                directions.add("D")
                outbound = outbound or flight_no
            internal_id = _text(mapped.get("safeguard_code"))
            if internal_id and internal_id not in internal_ids:
                internal_ids.append(internal_id)
            airline = airline or _text(mapped.get("airline"))
            aircraft_type = aircraft_type or _text(mapped.get("aircraft_type"))
            aircraft_no = aircraft_no or _text(mapped.get("aircraft_no"))
            raw_payloads.append(raw)
        if not directions:
            inbound = inbound or (flight_numbers[0] if flight_numbers else None)
            outbound = outbound or (flight_numbers[1] if len(flight_numbers) > 1 else None)
        return {
            "flight_key": f"{group_key}@{stand or 'UNKNOWN'}",
            "safeguard_code": internal_ids[0] if internal_ids else None,
            "inbound_flight_no": inbound,
            "outbound_flight_no": outbound,
            "stand": stand or None,
            "plan_start": min(starts) if starts else None,
            "plan_end": max(ends) if ends else None,
            "airline": airline,
            "aircraft_type": aircraft_type,
            "aircraft_no": aircraft_no,
            "issue_tags": [],
            "raw_payload": {
                "rows": raw_payloads,
                "internal_ids": internal_ids,
                "normalization": {"method": normalization_method},
                "directions": sorted(directions),
            },
        }


def _is_complementary_occupancy_pair(drafts: list[dict[str, Any]]) -> bool:
    if len(drafts) != 2:
        return False
    direction_sets = [set(draft["raw_payload"].get("directions", [])) for draft in drafts]
    return {frozenset(value) for value in direction_sets} == {
        frozenset({"A"}),
        frozenset({"D"}),
    }


def parse_nodes(content: bytes, source_type: str) -> list[dict[str, Any]]:
    output = []
    for raw in _read_rows(content):
        mapped = _alias_map(raw)
        message: dict[str, Any] = {}
        message_content = mapped.get("message_content")
        if isinstance(message_content, dict):
            message = message_content
        elif message_content:
            try:
                parsed = json.loads(str(message_content))
                message = _json_value(parsed) if isinstance(parsed, dict) else {}
            except (TypeError, ValueError, json.JSONDecodeError):
                message = {}

        if message.get("type") == "turnaround_flight":
            continue

        event_time = _datetime(mapped.get("event_time") or message.get("event_time"))
        stand = _text(mapped.get("stand") or message.get("stand_name"))
        event_type = _text(mapped.get("event_type") or message.get("event_name")) or "UnknownEvent"
        reported_flight = _text(
            mapped.get("flight_no")
            or mapped.get("inbound_flight_no")
            or mapped.get("outbound_flight_no")
            or message.get("flight_no")
        )
        raw_payload = dict(raw)
        if message:
            raw_payload["message_content"] = message
        output.append(
            {
                "source_type": source_type,
                "source_row_id": _text(mapped.get("source_row_id")) or str(raw["__row_number__"]),
                "event_type": event_type,
                "event_time": event_time,
                "stand": stand,
                "reported_flight_no": reported_flight,
                "safeguard_code": _text(mapped.get("safeguard_code")),
                "is_anomaly": event_time is None or stand is None,
                "raw_payload": raw_payload,
            }
        )
    return output


def source_summary(rows: Iterable[dict[str, Any]]) -> dict[str, int]:
    items = list(rows)
    return {
        "rows": len(items),
        "invalid": sum(bool(item.get("is_anomaly")) for item in items),
    }
